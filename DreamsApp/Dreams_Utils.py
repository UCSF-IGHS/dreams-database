from django.core.files.storage import default_storage
import openpyxl as xl
import xlrd
import traceback
from Dreams_Excel_Mapping import *
from string import Template
import datetime
from django.db import connection
from django.conf import settings
import os


class DreamsEnrollmentExcelDatabase(object):

    document_path = ''
    workbook = ''

    def __init__(self):
        pass

    def create_tmp_file(self, excel):
        try:
            with open(default_storage.path('tmp/enrollment.xlsx'), 'wb+') as destination:
                for chunk in excel.chunks():
                    destination.write(chunk)
            destination.closed
            self.document_path = default_storage.path('tmp/enrollment.xlsx')
            return default_storage.path('tmp/enrollment.xlsx')
        except:
            print 'Could not create a temp file'
            return

    def get_document_path(self):
        return self.document_path

    def validate_excel_document(self):
        # check to see if Main Database sheet exists
        # TODO define and implement more checks

        wb = xl.load_workbook(self.get_document_path(), read_only=True, keep_vba=False)
        self.workbook = wb
        if 'Main Database' in wb.get_sheet_names():
            return True
        else:
            return False

    def get_sheet_names(self):
        if self.validate_excel_document():
            return xl.load_workbook(self.get_document_path(), read_only=True, keep_vba=False).get_sheet_names()
        else:
            return 'Invalid Sheets'

    def excel_enrollment_data(self):
        if self.validate_excel_document():
            try:
                sql_statement = self.generate_insert_SQL(6, 50)
                # print sql_statement
                self.execute_SQL_Query(sql_statement)

            except Exception as e:
                tb = traceback.format_exc()
                print 'There was an error ' + tb
            return

    def generate_insert_SQL(self, start, end):
        if self.validate_excel_document():
            try:
                print 'The Excel Database is Valid'
                print "Now processing the file.........."

                sql_statement = """
INSERT INTO DreamsApp_client
(first_name, middle_name, last_name, dreams_id, dss_id_number, implementing_partner_id, date_of_birth, date_of_enrollment, marital_status_id, phone_number,
county_of_residence_id, sub_county_id, ward_id, informal_settlement, village, landmark, guardian_name, relationship_with_guardian, guardian_phone_number, guardian_national_id, is_date_of_birth_estimated, verification_document_id, verification_doc_no )
VALUES """

                rangeList = range(start, end)
                lastNumber = rangeList[-1]
                row_qry = ''
                for rowid in range(start, end):

                    if rowid != lastNumber:
                        row_qry += self.generate_row_insert_SQL(rowid) + ', \n'
                    else:
                        row_qry += " " + self.generate_row_insert_SQL(rowid)

                return Template('$insert $values').substitute(insert=sql_statement, values=row_qry)

            except Exception as e:
                tb = traceback.format_exc()
                print 'There was an error ' + tb
            return

    def get_enrollment_sheet_openpyxl(self):
        return xl.load_workbook(self.get_document_path(), read_only=True, data_only=True).get_sheet_by_name('Main Database')

    def get_enrollment_sheet(self):
        return self.get_wb().sheet_by_name('Main Database')

    def get_wb(self):
        return xlrd.open_workbook(self.get_document_path(), on_demand=True)

    def get_row_data(self, row_num):

        sheet = self.get_enrollment_sheet_openpyxl()
        cols = self.openpyxl_demographic_columns()

        # map fields that should mirror webapp's

        ver_doc = str(sheet.cell(row=row_num, column=cols.get('verification_doc')).value)
        ver_doc = ExcelDreamsMapping.verification_document().get(ver_doc)

        marital_status = str(sheet.cell(row=row_num, column=cols.get('marital_status')).value)
        marital_status = ExcelDreamsMapping.marital_status_codes().get(marital_status)

        county = str(sheet.cell(row=row_num, column=cols.get('county')).value)
        county = ExcelDreamsMapping.county().get(county)

        sub_county = str(sheet.cell(row=row_num, column=cols.get('sub_county')).value)
        sub_county = ExcelDreamsMapping.sub_county().get(sub_county)

        ward = str(sheet.cell(row=row_num, column=cols.get('ward')).value)
        ward = ExcelDreamsMapping.ward_by_name().get(ward)

        dob_raw = sheet.cell(row=row_num, column=cols.get('dob')).value
        if dob_raw is not None:
            dob_raw = datetime.datetime.strftime(dob_raw, '%Y-%m-%d')

        doe_raw = sheet.cell(row=row_num, column=cols.get('date_of_enrollment')).value
        if doe_raw is not None:
            doe_raw = datetime.datetime.strftime(doe_raw, '%Y-%m-%d')

        row_values = {
            'serial_No': str(sheet.cell(row=row_num, column=cols.get('serial_No')).value),
            'IP': str(sheet.cell(row=row_num, column=cols.get('IP')).value),
            'first_name': str(sheet.cell(row=row_num, column=cols.get('first_name')).value),
            'middle_name': str(sheet.cell(row=row_num, column=cols.get('middle_name')).value),
            'last_name': str(sheet.cell(row=row_num, column=cols.get('last_name')).value),
            'dob': dob_raw,
            'verification_doc': ver_doc,
            'verification_doc_other': str(sheet.cell(row=row_num, column=cols.get('verification_doc_other')).value),
            'verification_doc_no': str(sheet.cell(row=row_num, column=cols.get('verification_doc_no')).value),
            'date_of_enrollment': doe_raw,
            'marital_status': marital_status,
            'client_phone_no': str(sheet.cell(row=row_num, column=cols.get('client_phone_no')).value),
            'county': county,
            'sub_county': sub_county,
            'ward': ward,
            'informal_settlement': str(sheet.cell(row=row_num, column=cols.get('informal_settlement')).value),
            'village': str(sheet.cell(row=row_num, column=cols.get('village')).value),
            'land_mark': str(sheet.cell(row=row_num, column=cols.get('land_mark')).value),
            'dreams_id': str(sheet.cell(row=row_num, column=cols.get('dreams_id')).value),
            'dss_id': str(sheet.cell(row=row_num, column=cols.get('dss_id')).value),
            'caregiver_first_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_first_name')).value),
            'caregiver_middle_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_middle_name')).value),
            'caregiver_last_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_last_name')).value),
            'caregiver_relationship': str(sheet.cell(row=row_num, column=cols.get('caregiver_relationship')).value),
            'caregiver_relationship_other': str(sheet.cell(row=row_num, column=cols.get('caregiver_relationship_other')).value),
            'caregiver_phone_no': str(sheet.cell(row=row_num, column=cols.get('caregiver_phone_no')).value),
            'caregiver_id_no': str(sheet.cell(row=row_num, column=cols.get('caregiver_id_no')).value),
        }
        return row_values


    def generate_row_insert_SQL(self, row_num):

        row_data = self.get_row_data(row_num)
        caregiver_fname = row_data.get('caregiver_first_name')
        caregiver_mname = row_data.get('caregiver_middle_name')
        caregiver_lname = row_data.get('caregiver_last_name')

        coded_rel = row_data.get('caregiver_relationship')
        other_rel = row_data.get('caregiver_relationship_other')
        date_of_birth_estimated = 0

        # django app has no field for other relationship. check which one has value

        if coded_rel != 'None' or other_rel != 'None':
            if other_rel != 'None':
                caregiver_relationship = other_rel
            else:
                caregiver_relationship = coded_rel
        else:
            caregiver_relationship = ''

        ver_doc = row_data.get('verification_doc')
        ver_doc_other = row_data.get('verification_doc_other')

        # TODO: Handle other verification documents. It is not currently in the Client's model. Code therefore picks only coded documents

        guardian_name = (caregiver_fname + " " if caregiver_fname != 'None' else "") + (caregiver_mname + " " if caregiver_mname != 'None' else "") + (caregiver_lname if caregiver_lname != 'None' else "")

        values_template = Template(
            '("$first_name", "$middle_name", "$last_name", "$dreams_id", "$dss_id_number", "$implementing_partner_id", "$date_of_birth", "$date_of_enrollment", "$marital_status_id", "$phone_number", \
    "$county_of_residence_id", "$sub_county_id", "$ward_id", "$informal_settlement", "$village", "$landmark", "$guardian_name", "$relationship_with_guardian", "$guardian_phone_number", "$guardian_national_id", "$is_date_of_birth_estimated", "$verification_doc_id", "$verification_doc_no" )').safe_substitute(
            first_name=row_data.get('first_name') if row_data.get('first_name') !='None' else '',
            middle_name=row_data.get('middle_name') if row_data.get('middle_name') !='None' else '',
            last_name=row_data.get('last_name') if row_data.get('last_name') !='None' else '',
            dreams_id=row_data.get('dreams_id'),
            dss_id_number=row_data.get('dss_id') if row_data.get('dss_id') !='None' else '',
            implementing_partner_id=row_data.get('IP') if row_data.get('IP') !='None' else '',
            date_of_birth=row_data.get('dob') if row_data.get('dob') !='None' else '0000-00-00',
            date_of_enrollment=row_data.get('date_of_enrollment') if row_data.get('date_of_enrollment') !='None' else '0000-00-00',
            marital_status_id=row_data.get('marital_status') if row_data.get('marital_status') !='None' else '',
            phone_number=row_data.get('client_phone_no') if row_data.get('client_phone_no') !='None' else '',
            county_of_residence_id=row_data.get('county'),
            sub_county_id=row_data.get('sub_county'),
            ward_id=row_data.get('ward'),
            informal_settlement=row_data.get('informal_settlement') if row_data.get('informal_settlement') !='None' else '',
            village=row_data.get('village') if row_data.get('village') !='None' else '',
            landmark=row_data.get('land_mark') if row_data.get('land_mark') !='None' else '',
            guardian_name=guardian_name,
            relationship_with_guardian=caregiver_relationship,
            guardian_phone_number=row_data.get('caregiver_phone_no') if row_data.get('caregiver_phone_no') !='None' else '',
            guardian_national_id=row_data.get('caregiver_id_no') if row_data.get('caregiver_id_no') !='None' else '',
            is_date_of_birth_estimated=date_of_birth_estimated,
            verification_doc_id=row_data.get('verification_doc'),
            verification_doc_no=row_data.get('verification_doc_no')
        )

        return values_template

    def openpyxl_demographic_columns(self):
        demographics = {
            'serial_No': 1,
            'IP': 3,
            'first_name': 4,
            'middle_name': 5,
            'last_name': 6,
            'dob': 7,
            'verification_doc': 8,
            'verification_doc_other': 9,
            'verification_doc_no': 10,
            'date_of_enrollment': 11,
            'marital_status': 16,
            'client_phone_no': 17,
            'county': 18,
            'sub_county': 19,
            'ward': 20,
            'informal_settlement': 22,
            'village': 23,
            'land_mark': 24,
            'dreams_id': 25,
            'dss_id': 26,
            'caregiver_first_name': 27,
            'caregiver_middle_name': 28,
            'caregiver_last_name': 29,
            'caregiver_relationship': 30,
            'caregiver_relationship_other': 31,
            'caregiver_phone_no': 32,
            'caregiver_id_no': 33,
        }

        return demographics

    def dump_SQL(self):
        return """SELECT
d.first_name,d.middle_name,d.last_name,d.date_of_birth, d.verification_document_id,d.verification_doc_no,d.date_of_enrollment,d.phone_number,
  d.dss_id_number,d.informal_settlement,d.village,d.landmark,d.dreams_id,d.guardian_name,d.relationship_with_guardian,d.guardian_phone_number,
  d.guardian_national_id,d.date_created,d.county_of_residence_id,d.implementing_partner_id,d.marital_status_id,d.sub_county_id,d.ward_id,
i.head_of_household_id, i.head_of_household_other,i.age_of_household_head, i.is_father_alive, i.is_mother_alive, i.is_parent_chronically_ill,
  i.main_floor_material_id, i.main_floor_material_other, i.main_roof_material_id, i.main_roof_material_other, i.main_wall_material_id, i.main_wall_material_other,
  i.source_of_drinking_water_id,i.source_of_drinking_water_other, i.no_of_adults, i.no_of_females, i.no_of_males, i.no_of_children,
  i.currently_in_ct_program_id, i.current_ct_program, i.ever_enrolled_in_ct_program_id, i.ever_missed_full_day_food_in_4wks_id, i.has_disability_id,
  i.no_of_days_missed_food_in_4wks_id,
s.age_at_first_sexual_encounter,s.sex_partners_in_last_12months,s.age_of_last_partner_id,s.age_of_second_last_partner_id,
  s.age_of_third_last_partner_id,s.ever_had_sex_id,s.has_sexual_partner_id,s.know_last_partner_hiv_status_id,
  s.know_second_last_partner_hiv_status_id,s.know_third_last_partner_hiv_status_id,s.last_partner_circumcised_id,
  s.received_money_gift_for_sex_id,s.second_last_partner_circumcised_id,s.third_last_partner_circumcised_id,s.used_condom_with_last_partner_id,
  s.used_condom_with_second_last_partner_id,s.used_condom_with_third_last_partner_id,
rh.no_of_biological_children,rh.anc_facility_name,rh.known_fp_method_other,rh.current_fp_method_other,rh.reason_not_using_fp_other,
rh.current_anc_enrollment_id,rh.current_fp_method_id,rh.currently_pregnant_id,rh.currently_use_modern_fp_id,rh.fp_methods_awareness_id,
rh.has_biological_children_id,rh.reason_not_using_fp_id, rh.known_fp_methods,
dr.drug_abuse_last_12months_other,dr.drug_used_last_12months_other,dr.drug_abuse_last_12months_id,dr.frequency_of_alcohol_last_12months_id,
  dr.produced_alcohol_last_12months_id,dr.used_alcohol_last_12months_id , drugs_used_in_last_12_months,
 p.dreams_program_other,p.client_id , programmes_enrolled,
gbv.gbv_help_provider_other,gbv.preferred_gbv_help_provider_other,gbv.economic_threat_ever_id,gbv.economic_threat_last_3months_id,
gbv.humiliated_ever_id,gbv.humiliated_last_3months_id,gbv.insulted_ever_id,gbv.insulted_last_3months_id,
gbv.knowledge_of_gbv_help_centres_id,gbv.physical_violence_ever_id,gbv.physical_violence_last_3months_id,
gbv.physically_forced_other_sex_acts_ever_id,gbv.physically_forced_other_sex_acts_last_3months_id,
gbv.physically_forced_sex_ever_id,gbv.physically_forced_sex_last_3months_id,gbv.seek_help_after_gbv_id,
gbv.threatened_for_sexual_acts_ever_id,gbv.threatened_for_sexual_acts_last_3months_id,gbv.threats_to_hurt_ever_id,
gbv.threats_to_hurt_last_3months_id, providers_sought, preferred_providers,
edu.current_school_name,edu.current_class,edu.current_school_level_other,edu.current_education_supporter_other,
edu.reason_not_in_school_other,edu.dropout_class,edu.life_wish_other,edu.current_income_source_other,
edu.banking_place_other,edu.banking_place_id,edu.current_income_source_id,edu.current_school_level_id,
edu.current_school_type_id,edu.currently_in_school_id,edu.dropout_school_level_id,edu.has_savings_id,edu.last_time_in_school_id,
edu.life_wish_id,edu.reason_not_in_school_id, current_edu_supporter_list,
hiv.care_facility_enrolled,hiv.reason_not_in_hiv_care_other,hiv.reason_never_tested_for_hiv_other,hiv.enrolled_in_hiv_care_id,
hiv.ever_tested_for_hiv_id,hiv.knowledge_of_hiv_test_centres_id,hiv.last_test_result_id,hiv.period_last_tested_id,
hiv.reason_not_in_hiv_care_id, reason_not_tested_for_hiv
FROM
DreamsApp_client AS d
LEFT OUTER JOIN DreamsApp_clientindividualandhouseholddata i ON i.client_id = d.id
LEFT OUTER JOIN DreamsApp_clientsexualactivitydata s ON s.client_id = d.id
LEFT OUTER JOIN (
SELECT *
FROM DreamsApp_clientreproductivehealthdata rh
LEFT OUTER JOIN
(
SELECT
fp.clientreproductivehealthdata_id    AS rh_id,
group_concat(familyplanningmethod_id) AS known_fp_methods
FROM DreamsApp_clientreproductivehealthdata_known_fp_method fp
LEFT OUTER JOIN DreamsApp_clientreproductivehealthdata rh
ON fp.clientreproductivehealthdata_id = rh.id
GROUP BY fp.clientreproductivehealthdata_id
) fpm ON fpm.rh_id = rh.id) rh ON rh.client_id = d.id
LEFT OUTER JOIN (
SELECT *
FROM DreamsApp_clientdrugusedata dd
LEFT OUTER JOIN
(
SELECT
d.clientdrugusedata_id  AS dd_id,
group_concat(d.drug_id) AS drugs_used_in_last_12_months
FROM DreamsApp_clientdrugusedata_drug_used_last_12months d
LEFT OUTER JOIN DreamsApp_clientdrugusedata inner_dd ON d.clientdrugusedata_id = inner_dd.id
GROUP BY dd_id
) d ON d.dd_id = dd.id) dr ON dr.client_id = d.id
INNER JOIN (SELECT *
FROM DreamsApp_clientparticipationindreams pp
LEFT OUTER JOIN
(
SELECT
dp.clientparticipationindreams_id   AS pr_id,
group_concat(dp.dreamsprogramme_id) AS programmes_enrolled
FROM DreamsApp_clientparticipationindreams_dreams_program dp
LEFT OUTER JOIN DreamsApp_clientparticipationindreams inner_pp
ON dp.clientparticipationindreams_id = inner_pp.id
GROUP BY pr_id
) pr ON pr.pr_id = pp.id) p ON p.client_id = d.id
INNER JOIN (SELECT
gbv.*,
providers.provider_list   AS providers_sought,
p_providers.provider_list AS preferred_providers
FROM DreamsApp_clientgenderbasedviolencedata gbv
LEFT OUTER JOIN (
SELECT
  provider.clientgenderbasedviolencedata_id AS rec_id,
  group_concat(provider.gbvhelpprovider_id) AS provider_list
FROM DreamsApp_clientgenderbasedviolencedata_gbv_help_provider provider
GROUP BY rec_id
) providers ON providers.rec_id = gbv.id
LEFT OUTER JOIN (
SELECT
  provider.clientgenderbasedviolencedata_id AS rec_id,
  group_concat(provider.gbvhelpprovider_id) AS provider_list
FROM DreamsApp_clientgenderbasedviolencedata_preferred_gbv_help_p1bce provider -- DreamsApp_clientgenderbasedviolencedata_preferred_gbv_help_provider
GROUP BY rec_id
) p_providers ON p_providers.rec_id = gbv.id
GROUP BY gbv.id) gbv ON gbv.client_id = d.id
INNER JOIN (SELECT
ed.*,
edu_sup.current_edu_supporter_list
FROM DreamsApp_clienteducationandemploymentdata ed
LEFT OUTER JOIN (
SELECT
  s.clienteducationandemploymentdata_id AS rec_id,
  group_concat(educationsupporter_id)   AS current_edu_supporter_list
FROM DreamsApp_clienteducationandemploymentdata_current_educationebf4 s -- DreamsApp_clienteducationandemploymentdata_current_education_supporter s
GROUP BY rec_id
) edu_sup ON edu_sup.rec_id = ed.id
GROUP BY ed.id) edu ON edu.client_id = d.id
INNER JOIN (SELECT
hiv_d.*,
rn_not_tested.reason_not_tested_for_hiv
FROM DreamsApp_clienthivtestingdata hiv_d
LEFT OUTER JOIN (
SELECT
  rn.clienthivtestingdata_id                AS rec_id,
  group_concat(rn.reasonnottestedforhiv_id) AS reason_not_tested_for_hiv
FROM DreamsApp_clienthivtestingdata_reason_never_tested_for_hiv rn
GROUP BY rec_id
) rn_not_tested ON rn_not_tested.rec_id = hiv_d.id
GROUP BY hiv_d.id) hiv ON hiv.client_id = d.id
        """

    def execute_SQL_Query(self, sql):
        cursor = connection.cursor()
        try:
            cursor.execute(sql)

            print cursor.rowcount
        except Exception as e:
            print 'There was an Error running the query\n'
            traceback.format_exc()

        return

    def get_export_rows(self):
        sql = self.dump_SQL()
        cursor = connection.cursor()
        try:
            cursor.execute(sql)
            columns = [col[0] for col in cursor.description]
            return [
                dict(zip(columns, row))
                for row in cursor.fetchall()
                ]
        except Exception as e:
            print 'There was an Error running the query\n'
            traceback.format_exc()
        return

    def prepare_excel_doc(self):

        DREAMS_TEMPLATE = os.path.join(settings.BASE_DIR, 'templates/excel_template/dreams_export.xlsx')

        try:
            wb = xl.load_workbook(DREAMS_TEMPLATE)

            enrollment_sheet = wb.get_sheet_by_name("Enrollment")
            intervention_sheet = wb.get_sheet_by_name("Interventions")

            db_data = self.get_export_rows()
            i = 5
            for row in db_data:
                i += 1
                self.map_demographics(enrollment_sheet, i, row)
                self.map_individual_and_household(enrollment_sheet, i, row)
                self.map_sexuality(enrollment_sheet, i, row)
                self.map_reproductive_health(enrollment_sheet, i, row)
                self.map_drug_use(enrollment_sheet, i, row)
                self.map_education_and_employment(enrollment_sheet, i, row)
                self.map_gbv(enrollment_sheet, i, row)
                self.map_program_participation(enrollment_sheet, i, row)
                self.map_hiv_testing(enrollment_sheet, i, row)

            wb.save('dreams_enrollment_interventions.xlsx')
            print "Loading template successful"
            return wb
        except Exception as e:
            print 'There was an Error loading dreams template'
            traceback.format_exc()
        return

    def map_demographics(self, ws, i, row):
        cols = {
            'implementing_partner_id': 2,
            # 'IP_Code': 3,
            'first_name': 4,
            'middle_name': 5,
            'last_name': 6,
            'date_of_birth': 7,
            'verification_document_id': 8,
            # 'verification_doc_other': 9,
            'verification_doc_no': 10,
            'date_of_enrollment': 11,
            'marital_status_id': 14,
            'phone_number': 15,
            'county_of_residence_id': 16,
            'sub_county_id': 17,
            'ward_id': 18,
            # 'ward_code': 19,
            'informal_settlement': 20,
            'village': 21,
            'land_mark': 22,
            'dreams_id': 23,
            'dss_id_number': 24,
            # 'caregiver_first_name': 25,
            # 'caregiver_middle_name': 26,
            # 'caregiver_last_name': 27,
            'relationship_with_guardian': 28,
            # 'caregiver_relationship_other': 29,
            'guardian_phone_number': 30,
            'guardian_national_id': 31,
        }

        for k, v in cols.items():
            if k == 'implementing_partner_id':
                val = row.get(k)
                if val is not None:
                    partner = self.map_implementing_partner().get(val)
                    ws.cell(row=i, column=v, value=partner)
            elif k == 'verification_document_id':
                val = row.get(k)
                if val is not None:
                    item = self.map_verification_document().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'marital_status_id':
                val = row.get(k)
                if val is not None:
                    item = self.map_marital_status_codes().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_individual_and_household(self, ws, i, row):
        cols = {
            'head_of_household_id': 32,
            'head_of_household_other': 33,
            'age_of_household_head': 34,
            'is_father_alive': 35,
            'is_mother_alive': 36,
            'is_parent_chronically_ill': 37,
            'main_floor_material_id': 38,
            'main_floor_material_other': 39,
            'main_roof_material_id': 40,
            'main_roof_material_other': 41,
            'main_wall_material_id': 42,
            'main_wall_material_other': 43,
            'source_of_drinking_water_id': 44,
            'source_of_drinking_water_other': 45,
            'no_of_adults': 58,
            'no_of_females': 56,
            'no_of_males': 57,
            'no_of_children': 59,
            'currently_in_ct_program_id': 61,
            'current_ct_program': 62,
            'ever_enrolled_in_ct_program_id': 60,
            'ever_missed_full_day_food_in_4wks_id': 46,
            'has_disability_id': 48,
            'no_of_days_missed_food_in_4wks_id': 47
        }

        for k, v in cols.items():
            if k == 'head_of_household_id':
                val = row.get(k)
                if val is not None:
                    item = self.headOfHouseHoldDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'is_father_alive':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'is_mother_alive':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'is_parent_chronically_ill':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'main_floor_material_id':
                val = row.get(k)
                if val is not None:
                    item = self.floorMaterialDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'main_roof_material_id':
                val = row.get(k)
                if val is not None:
                    item = self.roofMaterialDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'main_wall_material_id':
                val = row.get(k)
                if val is not None:
                    item = self.wallMaterialDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'source_of_drinking_water_id':
                val = row.get(k)
                if val is not None:
                    item = self.drinkingWaterSourceDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'currently_in_ct_program_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'ever_enrolled_in_ct_program_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'ever_missed_full_day_food_in_4wks_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'no_of_days_missed_food_in_4wks_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_sexuality(self, ws, i, row):
        cols = {
            'age_at_first_sexual_encounter': 106,
            'sex_partners_in_last_12months': 107,
            'age_of_last_partner_id': 109,
            'age_of_second_last_partner_id': 110,
            'age_of_third_last_partner_id': 111,
            'ever_had_sex_id': 105,
            'has_sexual_partner_id': 108,
            'know_last_partner_hiv_status_id': 115,
            'know_second_last_partner_hiv_status_id': 116,
            'know_third_last_partner_hiv_status_id': 117,
            'last_partner_circumcised_id': 112,
            'received_money_gift_for_sex_id': 121,
            'second_last_partner_circumcised_id': 113,
            'third_last_partner_circumcised_id': 114,
            'used_condom_with_last_partner_id': 118,
            'used_condom_with_second_last_partner_id': 119,
            'used_condom_with_third_last_partner_id': 120
        }
        for k, v in cols.items():
            if k == 'age_of_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.relativeAgeDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'age_of_second_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.relativeAgeDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'age_of_third_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.relativeAgeDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'ever_had_sex_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'has_sexual_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'know_last_partner_hiv_status_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'know_second_last_partner_hiv_status_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'know_third_last_partner_hiv_status_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'last_partner_circumcised_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'second_last_partner_circumcised_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'third_last_partner_circumcised_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'received_money_gift_for_sex_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'used_condom_with_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'used_condom_with_second_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'used_condom_with_third_last_partner_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_reproductive_health(self, ws, i, row):
        cols = {
            'no_of_biological_children': 123,
            'anc_facility_name': 126,
            'known_fp_method_other': 134,
            'current_fp_method_other': 137,
            'reason_not_using_fp_other': 139,
            'current_anc_enrollment_id': 125,
            'current_fp_method_id': 136,
            'currently_pregnant_id': 124,
            'currently_use_modern_fp_id': 135,
            'fp_methods_awareness_id': 127,
            'has_biological_children_id': 122,
            'reason_not_using_fp_id': 138,
            'known_fp_methods': 128
        }
        for k, v in cols.items():
            if k == 'known_fp_methods':
                val = row.get(k)
                if val is not None:
                    fpm = val.split(",")
                    for m in fpm:
                        if m == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_fp_method().get(m), value='Yes')
            elif k == 'current_anc_enrollment_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'current_fp_method_id':
                val = row.get(k)
                if val is not None:
                    item = self.familyPlanningDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'currently_pregnant_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'currently_use_modern_fp_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'fp_methods_awareness_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'has_biological_children_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'reason_not_using_fp_id':
                val = row.get(k)
                if val is not None:
                    item = self.reasonNoInFPDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_drug_use(self, ws, i, row):
        cols = {
            # 'dr.drug_abuse_last_12months_other': 180,
            'drug_used_last_12months_other': 191,
            'drug_abuse_last_12months_id': 182,
            'frequency_of_alcohol_last_12months_id': 181,
            'produced_alcohol_last_12months_id': 192,
            'used_alcohol_last_12months_id': 180,
            'drugs_used_in_last_12_months': 183
        }
        for k, v in cols.items():
            if k == 'drugs_used_in_last_12_months':
                val = row.get(k)
                if val is not None:
                    drugs = val.split(",")
                    for d in drugs:
                        if d == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_drugs().get(d), value='Yes')
            elif k == 'frequency_of_alcohol_last_12months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'drug_abuse_last_12months_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'produced_alcohol_last_12months_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'used_alcohol_last_12months_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_program_participation(self, ws, i, row):
        cols = {
            'dreams_program_other': 201,
            'programmes_enrolled': 193
        }
        for k, v in cols.items():
            if k == 'programmes_enrolled':
                val = row.get(k)
                if val is not None:
                    pgs = val.split(",")
                    for p in pgs:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_dreams_programmes().get(p), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_gbv(self, ws, i, row):
        cols = {
            'gbv_help_provider_other': 167,
            'preferred_gbv_help_provider_other': 179,
            'economic_threat_ever_id': 146,
            'economic_threat_last_3months_id': 147,
            'humiliated_ever_id': 140,
            'humiliated_last_3months_id': 141,
            'insulted_ever_id': 144,
            'insulted_last_3months_id': 145,
            'knowledge_of_gbv_help_centres_id': 168,
            'physical_violence_ever_id': 148,
            'physical_violence_last_3months_id': 149,
            'physically_forced_other_sex_acts_ever_id': 152,
            'physically_forced_other_sex_acts_last_3months_id': 153,
            'physically_forced_sex_ever_id': 150,
            'physically_forced_sex_last_3months_id': 151,
            'seek_help_after_gbv_id': 156,
            'threatened_for_sexual_acts_ever_id': 154,
            'threatened_for_sexual_acts_last_3months_id': 155,
            'threats_to_hurt_ever_id': 142,
            'threats_to_hurt_last_3months_id': 143,
            'providers_sought': 157,
            'preferred_providers': 169,
        }
        for k, v in cols.items():
            if k == 'providers_sought':
                val = row.get(k)
                if val is not None:
                    sp = val.split(",")
                    for p in sp:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_gbv_sought_provider().get(p), value='Yes')
            elif k == 'preferred_providers':
                val = row.get(k)
                if val is not None:
                    sp = val.split(",")
                    for p in sp:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_gbv_preferred_provider().get(p), value='Yes')
            elif k == 'economic_threat_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'economic_threat_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'humiliated_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'humiliated_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'insulted_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'insulted_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'knowledge_of_gbv_help_centres_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physical_violence_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physical_violence_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physically_forced_other_sex_acts_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physically_forced_other_sex_acts_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physically_forced_sex_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'physically_forced_sex_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'seek_help_after_gbv_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'threatened_for_sexual_acts_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'threatened_for_sexual_acts_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'threats_to_hurt_ever_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'threats_to_hurt_last_3months_id':
                val = row.get(k)
                if val is not None:
                    item = self.frequencyDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_education_and_employment(self, ws, i, row):
        cols = {
            'current_school_name': 64,
            'current_class': 68,
            'current_school_level_other': 67,
            'current_edu_supporter_list': 69,
            'current_education_supporter_other': 70,
            'reason_not_in_school_other': 76,
            'dropout_class': 78,
            'life_wish_other': 81,
            'current_income_source_other': 83,
            'banking_place_other': 86,
            'banking_place_id': 85,
            'current_income_source_id': 82,
            'current_school_level_id': 66,
            'current_school_type_id': 65,
            'currently_in_school_id': 63,
            'dropout_school_level_id': 79,
            'has_savings_id': 84,
            'last_time_in_school_id': 77,
            'life_wish_id': 80,
            'reason_not_in_school_id': 75
        }
        for k, v in cols.items():
            if k == 'current_edu_supporter_list':
                val = row.get(k)
                if val is not None:
                    sps = val.split(",")
                    for s in sps:
                        if s == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_education_supporter().get(s), value='Yes')
            elif k == 'banking_place_id':
                val = row.get(k)
                if val is not None:
                    item = self.bankingPlaceDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'current_income_source_id':
                val = row.get(k)
                if val is not None:
                    item = self.incomeSourceDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'current_school_level_id':
                val = row.get(k)
                if val is not None:
                    item = self.schoolLevelDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'current_school_type_id':
                val = row.get(k)
                if val is not None:
                    item = self.schoolTypeDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'currently_in_school_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'dropout_school_level_id':
                val = row.get(k)
                if val is not None:
                    item = self.schoolLevelDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'has_savings_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'last_time_in_school_id':
                val = row.get(k)
                if val is not None:
                    item = self.lastInSchoolDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'life_wish_id':
                val = row.get(k)
                if val is not None:
                    item = self.lifeWishDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'reason_not_in_school_id':
                val = row.get(k)
                if val is not None:
                    item = self.reasonNotInSchoolDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_hiv_testing(self, ws, i, row):
        cols = {
            'care_facility_enrolled': 91,
            'reason_not_in_hiv_care_other': 93,
            'reason_not_tested_for_hiv': 94,
            'reason_never_tested_for_hiv_other': 103,
            'enrolled_in_hiv_care_id': 90,
            'ever_tested_for_hiv_id': 87,
            'knowledge_of_hiv_test_centres_id': 104,
            'last_test_result_id': 89,
            'period_last_tested_id': 88,
            'reason_not_in_hiv_care_id': 92
        }
        for k, v in cols.items():
            if k == 'reason_not_tested_for_hiv':
                val = row.get(k)
                if val is not None:
                    rns = val.split(",")
                    for r in rns:
                        if r == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_reason_never_tested_for_hiv().get(r), value='Yes')
            elif k == 'enrolled_in_hiv_care_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'ever_tested_for_hiv_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'knowledge_of_hiv_test_centres_id':
                val = row.get(k)
                if val is not None:
                    item = self.yesNoDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'last_test_result_id':
                val = row.get(k)
                if val is not None:
                    item = self.hivTestDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'period_last_tested_id':
                val = row.get(k)
                if val is not None:
                    item = self.periodDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            elif k == 'reason_not_in_hiv_care_id':
                val = row.get(k)
                if val is not None:
                    item = self.reasonNotInCareDictionary().get(val)
                    ws.cell(row=i, column=v, value=item)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_dreams_programmes(self):
        return {
            '1': 193,
            '2': 194,
            '3': 195,
            '4': 196,
            '5': 197,
            '6': 198,
            '7': 199,
            '8': 200,
            '96': 201
        }

    def map_drugs(self):
        return {
            '1': 183,
            '2': 184,
            '3': 185,
            '4': 186,
            '5': 187,
            '6': 188,
            '7': 189,
            '8': 190,
            '96': 191

        }

    def map_gbv_preferred_provider(self):
        return {
            '1': 169,
            '2': 170,
            '3': 171,
            '4': 172,
            '5': 173,
            '6': 174,
            '7': 175,
            '8': 176,
            '9': 177,
            '10': 178,
            '96': 179

        }

    def map_gbv_sought_provider(self):
        return {
            '1': 157,
            '2': 158,
            '3': 159,
            '4': 160,
            '5': 161,
            '6': 162,
            '7': 163,
            '8': 164,
            '9': 165,
            '10': 166,
            '96': 167

        }

    def map_fp_method(self):
        return {
            '1': 128,
            '2': 129,
            '3': 130,
            '4': 131,
            '5': 132,
            '6': 133,
            '96': 134

        }

    def map_reason_never_tested_for_hiv(self):
        return {
            '1': 94,
            '2': 95,
            '3': 96,
            '4': 97,
            '5': 98,
            '6': 99,
            '7': 100,
            '8': 101,
            '9': 102,
            '96': 103

        }

    def map_education_supporter(self):
        return {
            '1': 69,
            '2': 70,
            '3': 71,
            '4': 72,
            '5': 73,
            '96': 74
        }

    def map_implementing_partner(self):
        return {
            1: 'Afya Jijini',
            2: 'AIHA',
            3: 'Aphiaplus Western',
            4: 'Global Communities',
            5: 'Henry Jackson Foundation',
            6: 'HWWK',
            7: 'IMC',
            8: 'IRDO',
            9: 'LVCT Health',
            10: 'Peace Corps'
        }

    def map_verification_document(self):
        return {
            4: 'Baptismal card',
            1: 'Birth Certificate',
            2: 'National ID',
            3: 'National ID waiting card',
        }

    def map_marital_status_codes(self):
        return {
            1: 'Single',
            2: 'Married / Cohabiting',
            3: 'Separated / Divorced',
            4: 'Widowed',
        }




    def headOfHouseHoldDictionary(self):
        return {
            1: "Self",
            2: "Father",
            3: "Mother",
            3: "Sibling",
            5: "Uncle/Aunt",
            6: "Grandparents",
            7: "Husband/Partner",
            96: "Other/Specify"
        }

    def yesNoDictionary(self):
        return {
            1: "YES",
            2: "NO",
            3: "Don't Know",
            4: "Not Applicable"
        }

    def floorMaterialDictionary(self):
        return {
            1: "Earth/Mud/Dung/Sand",
            2: "Wood planks",
            3: "Ceramic tiles",
            4: "Cement",
            96: "Other (Specify)"
        }

    def wallMaterialDictionary(self):
        return {
            1: "No Walls",
            2: "Dung/Mud",
            3: "Stone with mud",
            4: "Plywood/Cardboard",
            5: "Carton",
            6: "Wood",
            7: "Stone/Cement",
            96: "Other (Specify)"
        }

    def roofMaterialDictionary(self):
        return {
            1: "Grass/Thatch/Makuti",
            2: "Tin cans",
            3: "Corrugated iron sheet",
            4: "Asbestos sheets",
            5: "Concrete",
            96: "Other (Specify)"
        }

    def drinkingWaterSourceDictionary(self):
        return {
            1: "Piped water",
            2: "Open well",
            3: "Covered well/borehole",
            4: "Surface water (River, Spring and Lakes)",
            5: "Rain water",
            96: "Other (Specify)"
        }

    def frequencyDictionary(self):
        return {
                    1: "Rarely (1-2 days)",
            2: "Sometimes (3-10 days)",
            3: "Often (more than 10 days)",
            4: "Always",
            5: "Sometimes",
            6: "Never",
            7: "Often",
            8: "Not in the last 3 months",
            9: "Everyday",
            10: "5 to 6 times a week",
            11: "3 to 4 times a week",
            12: "Once a week",
            13: "2 to 3 times a month",
            14: "Once a month",
            15: "3 to 11 times in the past year",
            16: "1 to 2 times in the past year",
            96: "Other (Specify)"

        }

    def schoolTypeDictionary(self):
        return {
            1: "Formal",
            2: "Informal"
        }

    def schoolLevelDictionary(self):
        return {
            1: "Nursery",
            2: "PRIMARY_LEVEL",
            3: "SECONDAY_LEVEL",
            4: "TERTIARY_LEVEL",
            5: "VOCATIONAL_LEVEL",
            96: "Other (Specify)"
        }


    def reasonNotInSchoolDictionary(self):
        return {
            1: "Completed High School",
            2: "Lack of fees",
            3: "Pregnancy",
            4: "Peer pressure",
            5: "Not interested",
            6: "Awaiting to join secondary",
            7: "Got married",
            96: "Other (Specify)"
        }

    def lastInSchoolDictionary(self):
        return {
            1: "Less than 6 months ago",
            2: "6 to 12 months",
            3: "1to 2 years ago",
            4: "More than 2 years",
            5: "Never attended school",
            96: "Other (Specify)"
        }

    def lifeWishDictionary(self):
        return {
            1: "Pursue a course",
            2: "Start a business",
            3: "Go back to school",
            4: "Get married",
            96: "Other (Specify)"
        }

    def incomeSourceDictionary(self):
        return {
            1: "Formally employed",
            2: "Business person",
            3: "Casual labour",
            4: "Petty trade",
            5: "Farmer",
            6: "None",
            96: "Other (Specify)"
        }

    def bankingPlaceDictionary(self):
        return {
            1: "At home",
            2: "Table banking",
            3: "In the bank",
            96: "Other (Specify)"
        }

    def periodDictionary(self):
        return {
            1: "Less than 6 Months ago",
            2: "6 - 12 months ago",
            3: "1- 2 years ago",
            4: "More than 2 years ago",
            5: "Never attended to school",
            6: "Less than 3 months ago",
            7: "3-5 months ago",
            8: "6-12 months ago",
            9: "More than 12 months ago"
        }

    def hivTestDictionary(self):
        return {
            1: "Positive",
            2: "Negative",
            3: "Don't Know",
            4: "Declined to disclose"
        }

    def reasonNotInCareDictionary(self):
        return {
            1: "Facility is too far away",
            2: "I don't know where clinic is",
            3: "I can't afford it",
            4: "I feel healthy/not sick",
            5: "I fear people will know I have HIV if I go to clinic",
            6: "I feel I will be discriminated against by people at a facility",
            7: "The providers at facility are unfriendly",
            8: "I am taking alternative medicine that is not availble at the clinic",
            9: "I,m too busy to go",
            96: "Other(specify)"
        }

    def relativeAgeDictionary(self):
        return {
            2: "Same Age",
            3: "Younger",
            4: "Older"
        }

    def familyPlanningDictionary(self):
        return {
            1: "Pills",
            2: "Injectable",
            3: "Implants",
            4: "IUCD",
            5: "Condom",
            6: "Permanent (Tube Ligation)",
            96: "Other (Specify)"
        }

    def reasonNoInFPDictionary(self):
        return {
           1: "Not sexually active",
           2: "Religious reasons",
           3: "Can not afford",
           4: "Do not know where to get",
           5: "Currently pregnant",
           96: "Other (Specify)"
        }