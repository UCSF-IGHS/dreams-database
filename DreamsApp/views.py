import os
import traceback

from django.contrib import messages
from django.core.urlresolvers import reverse
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponseServerError, HttpResponse
from django.core import serializers
from django.core.mail import EmailMessage
from django.core.exceptions import *
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import Group, Permission
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage
from django.db import connection as db_conn_2, transaction
import urllib.parse
import json
from datetime import date, timedelta, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font
from DreamsApp.Dreams_Utils_Plain import DreamsRawExportTemplateRenderer, settings
from DreamsApp.forms import *
from dateutil.relativedelta import relativedelta
from DreamsApp.service_layer import ClientEnrolmentServiceLayer, TransferServiceLayer, ReferralServiceLayer, FollowUpsServiceLayer


def get_enrollment_form_config_data(request):
    try:
        try:
            current_ip = request.user.implementingpartneruser.implementing_partner.code
        except Exception as e:
            current_ip = 0
        config_data = {
            'implementing_partners': ImplementingPartner.objects.all(),
            'verification_documents': VerificationDocument.objects.all(),
            'marital_status': MaritalStatus.objects.all(),
            'counties': County.objects.all(),
            'current_ip': current_ip
        }
        return config_data
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def log_custom_actions(user_id, table, row_id, action, search_text):
    try:
        audit = Audit()
        audit.user_id = user_id
        audit.table = table
        audit.row_id = row_id
        audit.action = action
        audit.search_text = search_text
        audit.save()
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def user_login(request):
    try:
        if request.user.is_authenticated():
            return redirect('clients')
        if request.method == 'GET':
            return render(request, 'login.html', {'page_title': 'Login', 'page': 'login'})
        elif request.method == 'POST':
            user_name = request.POST.get('inputUsername', '')
            pass_word = request.POST.get('inputPassword', '')

            audit = Audit()
            audit.user_id = 0
            audit.table = "DreamsApp_client"
            audit.row_id = 0
            audit.action = "LOGIN"

            if user_name == '' or pass_word == '':
                audit.search_text = "Missing login Credentials"
                audit.save()
                response_data = {
                    'status': 'fail',
                    'message': 'Missing username or password.',
                    'page': 'login',
                    'page_title': 'Login'
                }
            else:
                user = authenticate(username=user_name, password=pass_word)
                if user is not None:
                    if user.is_active:
                        login(request, user)
                        audit.user_id = user.id
                        audit.search_text = "Login Successful for username: " + user_name
                        audit.save()
                        response_data = {
                            'status': 'success',
                            'message': 'Login successfull'
                        }
                    else:
                        audit.search_text = "Failed login for username: " + user_name
                        audit.save()
                        response_data = {
                            'status': 'fail',
                            'message': 'Sorry, your account is disabled'
                        }
                else:
                    response_data = {
                        'status': 'fail',
                        'message': 'Incorrect username or password'
                    }
            return JsonResponse(json.dumps(response_data), safe=False)
        else:
            return render(request, 'login.html')
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def build_filter_client_queryset(turple_1, turple_2, turple_3, turple_4, turple_5, turple_6):
    try:
        return Client.objects.filter(Q(first_name__icontains=str(turple_1[0]), middle_name__icontains=str(turple_1[1]),
                                       last_name__icontains=str(turple_1[2])) |
                                     Q(first_name__icontains=str(turple_2[0]), middle_name__icontains=str(turple_2[1]),
                                       last_name__icontains=str(turple_2[2])) |
                                     Q(first_name__icontains=str(turple_3[0]), middle_name__icontains=str(turple_3[1]),
                                       last_name__icontains=str(turple_3[2])) |
                                     Q(first_name__icontains=str(turple_4[0]), middle_name__icontains=str(turple_4[1]),
                                       last_name__icontains=str(turple_4[2])) |
                                     Q(first_name__icontains=str(turple_5[0]), middle_name__icontains=str(turple_5[1]),
                                       last_name__icontains=str(turple_5[2])) |
                                     Q(first_name__icontains=str(turple_6[0]), middle_name__icontains=str(turple_6[1]),
                                       last_name__icontains=str(turple_6[2]))) \
            .exclude(voided=True) \
            .order_by('first_name') \
            .order_by('middle_name') \
            .order_by('last_name')
    except Exception as e:
        return Client.objects.all()[0]


def filter_clients(search_client_term, is_advanced_search, request):
    search_client_term_parts = search_client_term.split()
    # check number of parts
    parts_count = len(search_client_term_parts)
    if parts_count == 1:
        filter_text = search_client_term_parts[0]
        search_result = Client.objects.filter(Q(dreams_id__iexact=filter_text) |
                                              Q(first_name__iexact=filter_text) |
                                              Q(middle_name__iexact=filter_text) |
                                              Q(last_name__iexact=filter_text)) \
            .exclude(voided=True).order_by('first_name').order_by('middle_name').order_by('last_name')
    elif parts_count > 1:
        # this is not a dreams id search but a name search
        filter_text_1 = search_client_term_parts[0]
        filter_text_2 = search_client_term_parts[1]
        filter_text_3 = '' if parts_count == 2 else search_client_term_parts[2]
        # first name and middle name
        search_result = build_filter_client_queryset((filter_text_1, filter_text_2, filter_text_3),
                                                     (filter_text_2, filter_text_1, filter_text_3),
                                                     (filter_text_2, filter_text_3, filter_text_1),
                                                     (filter_text_3, filter_text_2, filter_text_1),
                                                     (filter_text_3, filter_text_1, filter_text_2),
                                                     (filter_text_1, filter_text_3, filter_text_2))
    if is_advanced_search == 'True':
        county_filter = str(
            request.GET.get('county', '') if request.method == 'GET' else request.POST.get('county', ''))
        if county_filter != '':
            search_result = search_result.filter(county_of_residence_id=int(county_filter))
        sub_county_filter = str(
            request.GET.get('sub_county', '') if request.method == 'GET' else request.POST.get('sub_county', ''))
        if sub_county_filter != '':
            search_result = search_result.filter(sub_county_id=int(sub_county_filter))
        ward_filter = str(request.GET.get('ward', '') if request.method == 'GET' else request.POST.get('ward', ''))
        if ward_filter != '':
            search_result = search_result.filter(ward_id=int(ward_filter))
        start_date_filter_original = request.GET.get('doe_start_filter',
                                                     '') if request.method == 'GET' else request.POST.get(
            'doe_start_filter', '')
        start_date_filter = '2015-10-01' if start_date_filter_original == '' else start_date_filter_original
        end_date_filter_original = request.GET.get('doe_end_filter',
                                                   dt.today()) if request.method == 'GET' else request.POST.get(
            'doe_end_filter', dt.today())
        end_date_filter = dt.today() if end_date_filter_original == '' else end_date_filter_original
        search_result = search_result.filter(date_of_enrollment__range=[start_date_filter, end_date_filter])  #
        return [search_result, is_advanced_search, county_filter, sub_county_filter, ward_filter,
                start_date_filter_original,
                end_date_filter_original]
    return [search_result, is_advanced_search, '', '', '', '', '']


def clients(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            # get search details -- search_client_term
            page = request.GET.get('page', 1) if request.method == 'GET' else request.POST.get('page', 1)
            is_advanced_search = request.GET.get('is_advanced_search',
                                                 'False') if request.method == 'GET' else request.POST.get(
                'is_advanced_search', 'False')
            search_client_term = request.GET.get('search_client_term',
                                                 '') if request.method == 'GET' else request.POST.get(
                'search_client_term', '')
            search_client_term = search_client_term.strip()
            if search_client_term != "":
                search_result_tuple = filter_clients(search_client_term, is_advanced_search, request)
                search_result = search_result_tuple[0]
                # check for permissions
                if not request.user.has_perm("DreamsApp.can_view_cross_ip_data"):
                    try:
                        ip = request.user.implementingpartneruser.implementing_partner

                        transfer_out = ClientTransfer.objects.filter(source_implementing_partner=ip).filter(
                            transfer_status=ClientTransferStatus.objects.get(
                                                      code__exact=TRANSFER_ACCEPTED_STATUS))

                        transfer_out_clients = search_result.filter(pk__in=transfer_out.values_list('client'))

                        search_result = search_result.filter(implementing_partner_id=ip)
                        search_result = search_result.union(transfer_out_clients)

                    except Exception as e:
                        search_result = Client.objects.all()[:0]
            else:
                search_result = Client.objects.all()[:0]
                search_result_tuple = [search_result, 'False', '', '', '', '', '']
            log_custom_actions(request.user.id, "DreamsApp_client", None, "SEARCH", search_client_term)
            try:
                current_ip = request.user.implementingpartneruser.implementing_partner.code
            except Exception as e:
                current_ip = 0
            if request.is_ajax():
                json_response = {
                    'search_result': serializers.serialize('json', search_result),
                    'search_client_term': search_client_term,
                    'can_manage_client': request.user.has_perm('auth.can_manage_client'),
                    'can_change_client': request.user.has_perm('auth.can_change_client'),
                    'can_delete_client': request.user.has_perm('auth.can_delete_client'),
                    'implementing_partners': ImplementingPartner.objects.all(),
                    'verification_documents': VerificationDocument.objects.all(),
                    'marital_status': MaritalStatus.objects.all(),
                    'counties': County.objects.all(),
                    'current_ip': current_ip,
                    'demo_form': DemographicsForm()
                }
                return JsonResponse(json_response, safe=False)
            else:
                # Non ajax request.. Do a paginator
                # do pagination
                try:
                    paginator = Paginator(search_result, 20)
                    client_paginator = paginator.page(page)
                except PageNotAnInteger:
                    client_paginator = paginator.page(1)  # Deliver the first page is page is not an integer
                except EmptyPage:
                    client_paginator = paginator.page(
                        paginator.num_pages)  # Deliver the last page if page is out of scope

                county_filter = search_result_tuple[2] if search_result_tuple[2] != '' else '0'
                sub_county_filter = search_result_tuple[3] if search_result_tuple[3] != '' else '0'
                sub_counties = SubCounty.objects.filter(county_id=int(county_filter))
                ward_filter = search_result_tuple[4] if search_result_tuple[4] != '' else '0'
                wards = Ward.objects.filter(sub_county_id=int(sub_county_filter))

                client_enrolment_service_layer = ClientEnrolmentServiceLayer(request.user)
                minimum_maximum_age = client_enrolment_service_layer.get_minimum_maximum_enrolment_age(client_enrolment_service_layer.ENROLMENT_CUTOFF_DATE)
                max_dob = datetime.now().date() - relativedelta(years=int(minimum_maximum_age[0]))
                min_dob = datetime.now().date() - relativedelta(years=int(minimum_maximum_age[1]))

                response_data = {
                    'page': 'clients',
                    'page_title': 'DREAMS Client List',
                    'search_client_term': search_client_term,
                    'client_paginator': client_paginator,
                    'status': 'success',
                    'implementing_partners': ImplementingPartner.objects.all(),
                    'verification_documents': VerificationDocument.objects.all(),
                    'marital_status': MaritalStatus.objects.all(),
                    'counties': County.objects.all(),
                    'current_ip': current_ip,
                    'demo_form': DemographicsForm(),
                    'is_advanced_search': search_result_tuple[1],
                    'county_filter': county_filter,
                    'sub_county_filter': sub_county_filter,
                    'sub_counties': sub_counties,
                    'ward_filter': ward_filter,
                    'wards': wards,
                    'start_date_filter': search_result_tuple[5],
                    'end_date_filter': search_result_tuple[6],
                    'max_dob': max_dob,
                    'min_dob': min_dob
                }
                # county_filter, sub_county_filter, ward_filter, start_date_filter, end_date_filter
                return render(request, 'clients.html', response_data)
        else:
            return redirect('login')
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def follow_ups(request):
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:
        client_id = request.GET.get('client_id', '') if request.method == 'GET' else request.POST.get('client_id', '')
        if client_id is not None and client_id != 0:
            try:
                client = Client.objects.get(id=client_id)
                client_follow_ups = ClientFollowUp.objects.filter(client=client)

                follow_up_service_layer = FollowUpsServiceLayer(request.user)
                follow_up_perms = {
                    'can_create_follow_up': follow_up_service_layer.can_create_followup(),
                    'can_delete_follow_up': follow_up_service_layer.can_delete_followup(),
                    'can_edit_follow_up': follow_up_service_layer.can_edit_followup(),
                    'can_view_follow_up': follow_up_service_layer.can_view_followup()
                }

                page = request.GET.get('page', 1)
                paginator = Paginator(client_follow_ups, 20)
                follow_up_types = ClientFollowUpType.objects.all()
                follow_up_result_types = ClientLTFUResultType.objects.all()

                try:
                    displayed_follow_ups = paginator.page(page)
                except PageNotAnInteger:
                    displayed_follow_ups = paginator.page(1)
                except EmptyPage:
                    displayed_follow_ups = paginator.page(paginator.num_pages)

                return render(request, 'client_follow_ups.html', {
                    'page': 'Follow Ups',
                    'page_title': 'Client Follow Ups Page',
                    'client': client,
                    'user': request.user,
                    'follow_up_perms': follow_up_perms,
                    'follow_ups': displayed_follow_ups,
                    'follow_up_types': follow_up_types,
                    'follow_up_result_types': follow_up_result_types
                })
            except Client.DoesNotExist:
                response_data = {
                    'status': 'failed',
                    'message': 'Operation not allowed. Client does not exist',
                    'client_id': client.id
                }
                return JsonResponse(json.dumps(response_data), safe=False)
            except Exception as e:
                return render(request, 'login.html')
    else:
        raise PermissionDenied


def client_profile(request):
    """ Returns client profile """
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:
        client_id = request.GET.get('client_id', '') if request.method == 'GET' else request.POST.get(
            'client_id', '')
        search_client_term = request.GET.get('search_client_term', '') if request.method == 'GET' else request.POST.get(
            'search_client_term', '')
        if client_id is not None and client_id != 0:
            try:
                ip = request.user.implementingpartneruser.implementing_partner
                if ip:
                    ip_code = ip.code
                else:
                    ip_code = None
            except Exception as e:
                ip_code = None

            try:
                client_found = Client.objects.get(id=client_id)
                is_editable_by_ip = client_found.is_editable_by_ip(ip)
                can_add_intervention = client_found.can_add_intervention(ip)
                client_status = client_found.get_client_status(ip)

                if client_found is not None:
                    # get cash transfer details
                    cash_transfer_details = ClientCashTransferDetails.objects.get(client=client_found)
                    # create cash transfer details form
                    cash_transfer_details_form = ClientCashTransferDetailsForm(instance=cash_transfer_details,
                                                                               current_AGYW=client_found)
                    cash_transfer_details_form.save(commit=False)

                return render(request, 'client_profile.html', {'page': 'clients',
                                                               'page_title': 'DREAMS Client Service Uptake',
                                                               'client': client_found,
                                                               'ct_form': cash_transfer_details_form,
                                                               'ct_id': cash_transfer_details.id,
                                                               'search_client_term': search_client_term,
                                                               'user': request.user,
                                                               'transfer_form': ClientTransferForm(ip_code=ip_code,
                                                                                                   initial={
                                                                                                       'client':
                                                                                                           client_found}),
                                                               'is_editable_by_ip': is_editable_by_ip,
                                                               'can_add_intervention': can_add_intervention,
                                                               'client_status': client_status
                                                               })
            except ClientCashTransferDetails.DoesNotExist:
                cash_transfer_details_form = ClientCashTransferDetailsForm(current_AGYW=client_found)
                return render(request, 'client_profile.html',
                              {'page': 'clients',
                               'page_title': 'DREAMS Client Service Uptake',
                               'client': client_found,
                               'ct_form': cash_transfer_details_form,
                               'search_client_term': search_client_term,
                               'user': request.user,
                               'transfer_form': ClientTransferForm(ip_code=ip_code, initial={'client': client_found}),
                               'is_editable_by_ip': is_editable_by_ip,
                               'can_add_intervention': can_add_intervention,
                               'client_status': client_status
                               })
            except Client.DoesNotExist:
                return render(request, 'login.html')
            except Exception as e:
                return render(request, 'login.html')
    else:
        raise PermissionDenied


def save_client(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if request.method == 'GET':
                return render(request, 'enrollment.html', {'client': None})
            elif request.method == 'POST' and request.is_ajax():
                # process saving user
                try:
                    ip_code = request.user.implementingpartneruser.implementing_partner.code
                except Exception as e:
                    response_data = {
                        'status': 'fail',
                        'message': 'Enrollment Failed. You do not belong to an implementing partner',
                        'client_id': None,
                        'can_manage_client': request.user.has_perm('auth.can_manage_client'),
                        'can_change_client': request.user.has_perm('auth.can_change_client'),
                        'can_delete_client': request.user.has_perm('auth.can_delete_client')
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)

                client_form = DemographicsForm(request.POST)

                if client_form.is_valid():
                    client_enrolment_service_layer = ClientEnrolmentServiceLayer(request.user)

                    if not client_enrolment_service_layer.is_within_enrolment_dates(client_form.cleaned_data['date_of_birth'], client_form.cleaned_data['date_of_enrollment']):
                        min_max_age = client_enrolment_service_layer.get_minimum_maximum_enrolment_age(client_enrolment_service_layer.ENROLMENT_CUTOFF_DATE)

                        response_data = {
                            'status': 'fail',
                            'message': "The client is not within the accepted age range. At the date of enrolment the age of the client must be between " + str(
                                min_max_age[0]) + " and " + str(min_max_age[1] + " years."),
                            'client_id': None,
                            'can_manage_client': request.user.has_perm('auth.can_manage_client'),
                            'can_change_client': request.user.has_perm('auth.can_change_client'),
                            'can_delete_client': request.user.has_perm('auth.can_delete_client')
                        }
                        return JsonResponse(json.dumps(response_data), safe=False)

                    client = client_form.save()

                    # Check client dreams_id
                    if client.dreams_id is None or not client.dreams_id:
                        # Generate client dreams_id
                        cursor = db_conn_2.cursor()
                        try:
                            cursor.execute(
                                """
                                SELECT (max(CONVERT(SUBSTRING_INDEX(dreams_id, '/', -1), UNSIGNED INTEGER )) + 1)
                                from DreamsApp_client WHERE dreams_id is not null and ward_id is not null
                                AND DreamsApp_client.implementing_partner_id=%s
                                AND DreamsApp_client.ward_id=%s AND DreamsApp_client.voided=0 group by implementing_partner_id, ward_id;""",
                                (ip_code, client.ward.id))
                            next_serial = cursor.fetchone()[0]
                            client.dreams_id = str(ip_code) + '/' + str(client.ward.code if client.ward != None else '') \
                                               + '/' + str(next_serial)
                        except Exception as e:
                            next_serial = 1
                            client.dreams_id = str(ip_code) + '/' + str(1) \
                                               + '/' + str(next_serial)
                        finally:
                            cursor.close()
                            client.save()
                    # cascade create other modules
                    ClientIndividualAndHouseholdData.objects.get_or_create(client=client)
                    ClientEducationAndEmploymentData.objects.get_or_create(client=client)
                    ClientHIVTestingData.objects.get_or_create(client=client)
                    ClientSexualActivityData.objects.get_or_create(client=client)
                    ClientReproductiveHealthData.objects.get_or_create(client=client)
                    ClientGenderBasedViolenceData.objects.get_or_create(client=client)
                    ClientDrugUseData.objects.get_or_create(client=client)
                    ClientParticipationInDreams.objects.get_or_create(client=client)
                    if request.is_ajax():
                        response_data = {
                            'status': 'success',
                            'message': 'Enrollment to DREAMS successful. Redirecting you to the full enrolment data view',
                            'client_id': client.id,
                            'can_manage_client': request.user.has_perm('auth.can_manage_client'),
                            'can_change_client': request.user.has_perm('auth.can_change_client'),
                            'can_delete_client': request.user.has_perm('auth.can_delete_client')
                        }
                        return JsonResponse(json.dumps(response_data), safe=False)
                    else:
                        # redirect to page
                        return redirect('clients')
                else:
                    response_data = {
                        'status': 'fail',
                        'message': client_form.errors,
                        'client_id': None,
                        'can_manage_client': request.user.has_perm('auth.can_manage_client'),
                        'can_change_client': request.user.has_perm('auth.can_change_client'),
                        'can_delete_client': request.user.has_perm('auth.can_delete_client')
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)
        else:
            raise PermissionDenied
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': str(e),
            'client_id': None,
            'can_manage_client': request.user.has_perm('auth.can_manage_client'),
            'can_change_client': request.user.has_perm('auth.can_change_client'),
            'can_delete_client': request.user.has_perm('auth.can_delete_client')
        }
        return JsonResponse(json.dumps(response_data), safe=False)


def edit_client(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:  # and request.user.is_superuser:
            if request.method == 'GET':
                client_id = int(request.GET['client_id'])
                client = Client.objects.defer('date_of_enrollment', 'date_of_birth').get(id__exact=client_id)
                if client is None:
                    redirect('clients')
                if request.is_ajax():
                    response_data = {'client': serializers.serialize('json', [client, ])}
                    return JsonResponse(response_data, safe=False)
                else:
                    # redirect to page
                    return render(request, 'enrollment.html', {'client': client})
                return redirect('clients')
            elif request.method == 'POST':
                client_id = int(str(request.POST.get('client_id')))
                client = Client.objects.filter(id=client_id).first()

                if not client.is_editable_by_ip(request.user.implementingpartneruser.implementing_partner):
                    response_data = {
                        'status': 'failed',
                        'message': 'Operation not allowed. Client is not editable by your Implementing partner',
                        'client_id': client.id
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)

                if client.implementing_partner != request.user.implementingpartneruser.implementing_partner:
                    # user and client IPs dont match. Return error message
                    response_data = {
                        'status': 'failed',
                        'message': 'Operation not allowed. Client is not enrolled by your Implementing partner',
                        'client_id': client.id
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)

                client_enrolment_service_layer = ClientEnrolmentServiceLayer(request.user)
                date_of_birth = datetime.strptime(request.POST.get('date_of_birth'), '%Y-%m-%d').date()
                date_of_enrollment = datetime.strptime(request.POST.get('date_of_enrollment'), '%Y-%m-%d').date()

                if not client_enrolment_service_layer.is_within_enrolment_dates(date_of_birth, date_of_enrollment):
                    min_max_age = client_enrolment_service_layer.get_minimum_maximum_enrolment_age(
                        client_enrolment_service_layer.ENROLMENT_CUTOFF_DATE)

                    response_data = {
                        'status': 'failed',
                        'message': "The client is not within the accepted age range. At the date of enrolment the age of the client must be between " + str(
                            min_max_age[0]) + " and " + str(min_max_age[1] + " years."),
                        'client_id': client.id
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)

                # process editing user
                client.implementing_partner = ImplementingPartner.objects.filter(
                    code__exact=str(request.POST.get('implementing_partner', ''))).first()
                client.first_name = str(request.POST.get('first_name', ''))
                client.middle_name = str(request.POST.get('middle_name', ''))
                client.last_name = str(request.POST.get('last_name', ''))
                client.date_of_birth = str(date_of_birth)
                client.is_date_of_birth_estimated = bool(str(request.POST.get('is_date_of_birth_estimated')))
                client.verification_document = VerificationDocument.objects.filter(
                    code__exact=str(request.POST.get('verification_document', ''))).first()
                client.verification_doc_no = str(request.POST.get('verification_doc_no', ''))
                client.date_of_enrollment = str(datetime.strptime(request.POST.get('date_of_enrollment', dt.now()), '%Y-%m-%d').date())
                client.age_at_enrollment = int(str(request.POST.get('age_at_enrollment')))
                client.marital_status = MaritalStatus.objects.filter(
                    code__exact=str(request.POST.get('marital_status', ''))).first()
                client.phone_number = str(request.POST.get('phone_number', ''))
                client.dss_id_number = str(request.POST.get('dss_id_number', ''))
                client.county_of_residence = County.objects.filter(
                    code__exact=request.POST.get('county_of_residence', '')).first()
                client.sub_county = SubCounty.objects.filter(code__exact=request.POST.get('sub_county', '')).first()
                client.ward = Ward.objects.filter(code__exact=request.POST.get('ward', 0)).first()
                client.informal_settlement = str(request.POST.get('informal_settlement', ''))
                client.village = str(request.POST.get('village', ''))
                client.landmark = str(request.POST.get('landmark', ''))
                client.dreams_id = str(request.POST.get('dreams_id', ''))
                client.guardian_name = str(request.POST.get('guardian_name', ''))
                client.relationship_with_guardian = str(request.POST.get('relationship_with_guardian', ''))
                client.guardian_phone_number = str(request.POST.get('guardian_phone_number', ''))
                client.guardian_national_id = str(request.POST.get('guardian_national_id', ''))
                client.save(user_id=request.user.id, action="UPDATE")
                if request.is_ajax():
                    response_data = {
                        'status': 'success',
                        'message': 'Client Details Updated successfuly.',
                        'client_id': client.id,
                        'can_manage_client': request.user.has_perm('auth.can_manage_client'),
                        'can_change_client': request.user.has_perm('auth.can_change_client'),
                        'can_delete_client': request.user.has_perm('auth.can_delete_client')
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)
                else:
                    # redirect to page
                    return redirect('clients')

        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def delete_client(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if request.method == 'GET' and request.is_ajax():
                client_id = int(request.GET['client_id'])
                client = Client.objects.filter(id__exact=client_id).first()

                if not client.is_editable_by_ip(request.user.implementingpartneruser.implementing_partner):
                    response_data = {
                        'status': 'failed',
                        'message': 'Operation not allowed. Client is not editable by your Implementing partner',
                        'client_id': client.id
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)

                # check if client and user IPs match
                if client.implementing_partner != request.user.implementingpartneruser.implementing_partner:
                    response_data = {
                        'status': 'failed',
                        'message': 'Operation not allowed. Client is not enrolled by your Implementing partner',
                        'client_id': client.id
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)

                # check if client has interventions
                if Intervention.objects.filter(client=client).count() > 0:
                    # Upating audit log
                    log_custom_actions(request.user.id, "DreamsApp_client", client_id, "DELETE", 'FAILED')
                    response_data = {
                        'status': 'fail',
                        'message': 'This client cannot be deleted because they have interventions.'
                    }
                else:
                    client.delete()
                    # Upating audit log
                    log_custom_actions(request.user.id, "DreamsApp_client", client_id, "DELETE", 'SUCCESS')
                    response_data = {
                        'status': 'success',
                        'message': 'Client Details Deleted successfuly.'
                    }
                return JsonResponse(json.dumps(response_data), safe=False)

            elif request.method == 'POST':
                raise PermissionDenied
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def get_client_status(client):
    status = ''
    try:
        if client.voided:
            status += ' Voided'
        if client.exited:
            if status != '':
                status += ' & '
            status += 'Exited'
        if status != '':
            status = status[:0] + '( ' + status[0:]
            last_index = len(status)
            status = status[:last_index] + ' ) ' + status[last_index:]
        return status
    except Exception as e:
        return 'Invalid Status'


def is_not_null_or_empty(str):
    return str is not None and str is not ""


def unexit_client(request):
    if request.user is not None and request.user.is_authenticated() and request.user.is_active and request.user.has_perm(
            'DreamsApp.can_exit_client'):
        try:
            client_id = int(str(request.POST.get('client_id', '0')))
            reason_for_exit = str(request.POST.get('reason_for_unexit', ''))
            date_of_exit = request.POST.get('date_of_unexit', datetime.now())
            client = Client.objects.filter(id=client_id).first()
            client.exited = not client.exited
            client.reason_exited = reason_for_exit
            client.exited_by = request.user
            client.date_exited = date_of_exit
            client.save()
            response_data = {
                'status': 'success',
                'message': 'Client Exit Undone',
                'client_id': client.id,
                'client_status': get_client_status(client)
            }
            return JsonResponse(response_data, status=200)
        except Exception as e:
            response_data = {
                'status': 'failed',
                'message': 'Invalid client Id: ' + str(e)
            }
            return JsonResponse(response_data, status=500)
    else:
        response_data = {
            'status': 'failed',
            'message': 'Permission Denied. Please contact System Administrator for help.'
        }
        return JsonResponse(response_data, status=500)


def exit_client(request):
    OTHER_CODE = 6

    if request.user is not None and request.user.is_authenticated() and request.user.is_active and request.user.has_perm(
            'DreamsApp.can_exit_client'):
        try:
            client_id = int(str(request.POST.get('client_id')))
            reason_for_exit = ExitReason.objects.get(id__exact=int(request.POST.get('reason_for_exit', '')))
            date_of_exit = request.POST.get('date_of_exit', datetime.now())
            exit_comment = request.POST.get('exit_comment')

            if reason_for_exit is not None:
                if reason_for_exit.code == OTHER_CODE:
                    if is_not_null_or_empty(exit_comment):
                        exited_client = other_client_exit(client_id, reason_for_exit, exit_comment, request.user, date_of_exit)
                    else:
                        raise Exception('Reason for exit missing')
                else:
                    exited_client = client_exit(client_id, reason_for_exit, request.user, date_of_exit)

            response_data = {
                'status': 'success',
                'message': 'Client Exited',
                'client_id': exited_client.id,
                'client_status': get_client_status(exited_client)
            }
            return JsonResponse(response_data, status=200)
        except Exception as e:
            response_data = {
                'status': 'failed',
                'message': 'Invalid client Id: ' + str(e)
            }
            return JsonResponse(response_data, status=500)
    else:
        response_data = {
            'status': 'failed',
            'message': 'Permission Denied. Please contact System Administrator for help.'
        }
        return JsonResponse(response_data, status=500)


def other_client_exit(client_id, reason_for_exit, exit_comment, exit_user, date_of_exit):
    client = Client.objects.filter(id=client_id).first()
    client.exited = True
    client.exit_reason = reason_for_exit
    client.reason_exited = exit_comment
    client.exited_by = exit_user
    client.date_exited = date_of_exit
    client.save()
    return client


def client_exit(client_id, reason_for_exit, exit_user, date_of_exit):
    client = Client.objects.filter(id=client_id).first()
    client.exited = True
    client.exit_reason = reason_for_exit
    client.exited_by = exit_user
    client.date_exited = date_of_exit
    client.save()
    return client


def testajax(request):
    return render(request, 'testAjax.html')


# Use /ivgetTypes/ in the post url to access the method
# Handles post request for intervention types.
# Receives category_code from request and searches for types in the database

def get_external_organisation(request):
    try:
        if request.method == 'GET' and request.user is not None and request.user.is_authenticated() and request.user.is_active:
            response_data = {}
            external_orgs = serializers.serialize('json', ExternalOrganisation.objects.all())
            response_data["external_orgs"] = external_orgs
            return JsonResponse(response_data)
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def get_unsuccessful_followup_attempts(request):
    try:
        if is_valid_get_request(request):
            response_data = {}
            client_id = int(request.GET.get('current_client_id'))
            client = Client.objects.get(id=client_id)
            unsuccessful_follow_up_attempts = ClientFollowUp.objects.filter(client=client,
                                                                            result_of_followup=ClientLTFUResultType.objects.filter(name='Lost').first()).all()
            response_data['unsuccessful_follow_up_attempts'] = len(unsuccessful_follow_up_attempts)
            return JsonResponse(response_data)
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def get_exit_reasons(request):
    try:
        if is_valid_get_request(request):
            response_data = {}
            exit_reasons = serializers.serialize('json', ExitReason.objects.all())
            response_data["exit_reasons"] = exit_reasons
            return JsonResponse(response_data)
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def is_valid_get_request(request):
    return request.method == 'GET' and request.user is not None and request.user.is_authenticated() and request.user.is_active

def is_valid_post_request(request):
    return request.method == 'POST' and request.user is not None and request.user.is_authenticated() and request.user.is_active

def get_intervention_types(request):
    try:
        if request.method == 'POST' and request.user is not None and request.user.is_authenticated() and request.user.is_active:
            response_data = {}
            category_code = request.POST.get('category_code')
            current_client_id = request.POST.get('current_client_id', 0)
            # get current client
            current_client = Client.objects.filter(id__exact=current_client_id).first()
            if current_client is None:
                raise Exception
            # Get category by code and gets all related types
            # Returns an object with itypes property
            given_intervention_type_ids = Intervention.objects.values_list('intervention_type', flat=True). \
                filter(client=current_client). \
                distinct()  # select distinct intervention type ids given to a user
            i_category = InterventionCategory.objects.get(code__exact=category_code)
            # compute age at enrollment
            current_age = current_client.get_current_age()
            i_types = InterventionType.objects.filter(intervention_category__exact=i_category.id, ) \
                .order_by('code')
            # .exclude(is_given_once=True, id__in=given_intervention_type_ids).order_by('code')
            """This code has been commented out to allow for change of intervention types for one time interventions"""
            # get id's of interventions that can only be given once and are already given
            i_types = serializers.serialize('json', i_types)
            response_data["itypes"] = i_types
            return JsonResponse(response_data)
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


# use /ivSave/ to post to the method
# Gets intervention_type_id,  from request


def save_intervention(request):
    try:
        if request.method == 'POST' and request.user is not None and request.user.is_authenticated() \
                and request.user.is_active and request.user.has_perm('DreamsApp.add_intervention'):
            try:
                client = Client.objects.get(id__exact=int(request.POST.get('client')))
                status = True
                if client.voided:
                    message = 'Error: You cannot Add Sevices to a voided Client. ' \
                              'Please contact System Administrator for help.'
                    status = False
                if client.exited:
                    message = 'Error: You cannot Add Sevices to a Client Exited from DREAMS. ' \
                              'Please contact System Administrator for help.'
                    status = False
                if not status:
                    response_data = {
                        'status': 'fail',
                        'message': message
                    }
                    return JsonResponse(response_data)
            except:
                response_data = {
                    'status': 'fail',
                    'message': "Error: Client with given ID Cannot be found. "
                               "Please contact System Administrator for help."
                }
                return JsonResponse(response_data)
            # check
            # Check if user belongs to an Ip
            if request.user.implementingpartneruser.implementing_partner is not None:
                intervention_type_code = int(request.POST.get('intervention_type_code'))
                intervention_type = InterventionType.objects.get(code__exact=intervention_type_code)
                """Check that this is not a one time intervention that has already been given to the client"""
                try:
                    """Get client intervention filtered by intervention types"""
                    client_interventions = Intervention.objects.filter(intervention_type=intervention_type,
                                                                       client=client).exclude(voided=True)
                    client_interventions_count = client_interventions.count()
                    if intervention_type.is_given_once and client_interventions.count() > 0:
                        """An intervention has been found. This is an error
                            Return error message to user
                        """
                        response_data = {
                            'status': 'fail',
                            'message': "Error: This is a one time service that has already been offered. Please conside"
                                       "r editing if necessary"
                        }
                        return JsonResponse(response_data)
                except Exception as e:
                    """An error has occurred. Throw exception. This will be handled elsewhere"""
                    raise Exception(str(e))

                intervention_date = dt.strptime(request.POST.get('intervention_date'), '%Y-%m-%d').date()

                # check if external organisation is selected
                external_organization_checkbox = request.POST.get('external_organization_checkbox')
                external_organization_code = request.POST.get('external_organization_code')
                other_external_organization_code = request.POST.get('other_external_organization_code')

                if external_organization_checkbox:
                    if not external_organization_code:
                        response_data = {
                            'status': 'fail',
                            'message': "Error: External organisation must be selected if checkbox is checked."
                        }
                        return JsonResponse(response_data)
                else:
                    if client.date_of_enrollment is not None and intervention_date < client.date_of_enrollment:
                        response_data = {
                            'status': 'fail',
                            'message': "Error: The intervention date must be after the client's enrollment date. "
                        }
                        return JsonResponse(response_data)

                if intervention_date > dt.now().date():
                    response_data = {
                        'status': 'fail',
                        'message': "Error: The intervention date must be before or on the current date. "
                    }
                    return JsonResponse(response_data)

                if intervention_type_code is not None and type(intervention_type_code) is int:
                    intervention = Intervention()
                    intervention.client = client
                    intervention.intervention_type = intervention_type
                    intervention.name_specified = request.POST.get('other_specify',
                                                                   '') if intervention_type.is_specified else ''
                    intervention.intervention_date = request.POST.get('intervention_date')
                    created_by = User.objects.get(id__exact=int(request.POST.get('created_by')))
                    intervention.created_by = created_by
                    intervention.date_created = dt.now()
                    intervention.comment = request.POST.get('comment', '')

                    if external_organization_checkbox:
                        intervention.external_organisation = ExternalOrganisation.objects.get(pk=external_organization_code)
                        if other_external_organization_code:
                            intervention.external_organisation_other = other_external_organization_code
                        else:
                            intervention.external_organisation_other = None

                    if intervention_type.has_hts_result:
                        intervention.hts_result = HTSResult.objects.get(code__exact=int(request.POST.get('hts_result')))

                    if intervention_type.has_pregnancy_result:
                        intervention.pregnancy_test_result = PregnancyTestResult.objects.get(
                            code__exact=int(request.POST.get('pregnancy_test_result')))

                    if intervention_type.has_ccc_number:
                        intervention.client_ccc_number = request.POST.get('client_ccc_number')

                    if intervention_type.has_no_of_sessions:
                        intervention.no_of_sessions_attended = request.POST.get('no_of_sessions_attended')

                    # Update implementing Partner
                    intervention.implementing_partner = ImplementingPartner.objects. \
                        get(id__exact=created_by.implementingpartneruser.implementing_partner.id)
                    intervention.save(user_id=request.user.id, action="INSERT")  # Logging
                    # using defer() miraculously solved serialization problem of datetime properties.
                    intervention = Intervention.objects.defer('date_changed', 'intervention_date', 'date_created'). \
                        get(id__exact=intervention.id)

                    is_editable_by_ip = {}
                    is_editable_by_ip[intervention.pk] = intervention.is_editable_by_ip(
                        request.user.implementingpartneruser.implementing_partner)

                    is_visible_by_ip = {}
                    is_visible_by_ip[intervention.pk] = intervention.is_visible_by_ip(
                        request.user.implementingpartneruser.implementing_partner)

                    response_data = {
                        'status': 'success',
                        'message': 'Intervention successfully saved',
                        'intervention': serializers.serialize('json', [intervention, ], ensure_ascii=False),
                        'i_type': serializers.serialize('json', [intervention_type]),
                        'hts_results': serializers.serialize('json', HTSResult.objects.all()),
                        'pregnancy_results': serializers.serialize('json', PregnancyTestResult.objects.all()),
                        'permissions': json.dumps({
                            'can_change_intervention': request.user.has_perm('DreamsApp.change_intervention'),
                            'can_delete_intervention': request.user.has_perm('DreamsApp.delete_intervention')
                        }),
                        'is_editable_by_ip': is_editable_by_ip,
                        'is_visible_by_ip': is_visible_by_ip
                    }
                    return JsonResponse(response_data)
                else:  # Invalid Intervention Type
                    response_data = {
                        'status': 'fail',
                        'message': "Error: Invalid Intervention Type. "
                                   "Please select a valid Intervention Type to Proceed"
                    }
                return JsonResponse(response_data)
            else:  # User has no valid IP.
                response_data = {
                    'status': 'fail',
                    'message': "Error: You do not belong to an Implementing Partner. "
                               "Please contact your system admin to add you to the relevant Implementing Partner."
                }
                return JsonResponse(response_data)
        else:
            response_data = {
                'status': 'fail',
                'message': "Permission Denied: You don't have permission to Add Intervention"
            }
            return JsonResponse(response_data)
    except ImplementingPartnerUser.DoesNotExist:
        response_data = {
            'status': 'fail',
            'message': "Error: You do not belong to an Implementing Partner. "
                       "Please contact your system admin to add you to the relevant Implementing Partner."
        }
        return JsonResponse(response_data)
    except Exception as e:
        # check if validation error
        if type(e) is ValidationError:
            errormsg = '; '.join(ValidationError(e).messages)
        else:
            errormsg = str(e)

        response_data = {
            'status': 'fail',
            'message': errormsg
        }
        return JsonResponse(response_data)


# method that returns a list of interventions of given category for a given client
# use /ivList/ url pattern to access the method


def get_intervention_list(request):
    try:
        if request.method == 'POST' and request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if 'client_id' not in request.POST or request.POST.get('client_id') == 0:
                return ValueError('No Client id found in your request! Ensure it is provided')
            if 'intervention_category_code' not in request.POST or request.POST.get('intervention_category_code') == 0:
                return ValueError('No Intervention Category Code found in your request! Ensure it is provided')

            client_id = request.POST.get('client_id')
            intervention_category_code = request.POST.get('intervention_category_code')
            iv_category = InterventionCategory.objects.get(code__exact=intervention_category_code)
            list_of_related_iv_types = InterventionType.objects.filter(intervention_category__exact=iv_category)
            iv_type_ids = [i_type.id for i_type in list_of_related_iv_types]
            # check for see_other_ip_data persmission

            list_of_interventions = Intervention.objects.defer('date_changed', 'intervention_date',
                                                               'date_created').filter(client__exact=client_id,
                                                                                      intervention_type__in=iv_type_ids,
                                                                                      voided=False) \
                .order_by('-intervention_date', '-date_created', '-date_changed')

            client_found = Client.objects.get(id=client_id)
            client_is_transferred_out = client_found.transferred_out(request.user.implementingpartneruser.implementing_partner)

            if not request.user.has_perm('DreamsApp.can_view_cross_ip_data'):
                if client_is_transferred_out:
                    list_of_interventions = list_of_interventions.filter(
                        implementing_partner_id=request.user.implementingpartneruser.implementing_partner.id)

            if not request.user.has_perm('auth.can_view_older_records'):
                list_of_interventions = list_of_interventions.filter(date_created__range=
                                                                     [dt.now() - timedelta(days=31),
                                                                      dt.now()]
                                                                     )

            is_editable_by_ip = {}
            is_visible_by_ip = {}
            for i in list_of_interventions:
                is_editable_by_ip[i.pk] = i.is_editable_by_ip(request.user.implementingpartneruser.implementing_partner)
                is_visible_by_ip[i.pk] = i.is_visible_by_ip(request.user.implementingpartneruser.implementing_partner)

            response_data = {
                'iv_types': serializers.serialize('json', list_of_related_iv_types),
                'interventions': serializers.serialize('json', list_of_interventions),
                'hts_results': serializers.serialize('json', HTSResult.objects.all()),
                'pregnancy_results': serializers.serialize('json', PregnancyTestResult.objects.all()),
                'permissions': json.dumps({
                    'can_change_intervention': request.user.has_perm('DreamsApp.change_intervention'),
                    'can_delete_intervention': request.user.has_perm('DreamsApp.delete_intervention')
                }),
                'is_editable_by_ip': is_editable_by_ip,
                'is_visible_by_ip': is_visible_by_ip
            }
            return JsonResponse(response_data)
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


# Gets an intervention. Takes intervention_id and returns Intervention object
# use /ivGet/ to access this method


def get_intervention(request):
    try:
        if request.method == 'POST' and request.user is not None and request.user.is_authenticated() and request.user.is_active:
            intervention_id = request.POST.get('intervention_id')
            if 'intervention_id' not in request.POST:
                return ValueError('No intervention id found in your request!')
            intervention = Intervention.objects.defer('date_changed', 'intervention_date', 'date_created').get(
                id__exact=intervention_id)
            if intervention is not None:
                response_data = {'intervention': serializers.serialize('json', [intervention, ])}
                return JsonResponse(response_data)
            else:
                raise Exception
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def add_follow_up(request):
    try:
        if is_valid_post_request(request):
            client_id = int(request.POST['client'], 0)
            client = Client.objects.get(id=client_id)

            follow_up_type = ClientFollowUpType.objects.filter(id__exact=request.POST.get('follow_up_type')).first()
            follow_up_result_type = ClientLTFUResultType.objects.filter(id__exact=request.POST.get('follow_up_result_type')).first()
            follow_up_date = request.POST.get('follow_up_date')
            follow_up_comments = request.POST.get('follow_up_comments')

            follow_up = ClientFollowUp()
            follow_up.client = client
            follow_up.date_of_followup = follow_up_date
            follow_up.type_of_followup = follow_up_type
            follow_up.result_of_followup = follow_up_result_type
            follow_up.comment = follow_up_comments
            follow_up.save()

            response_data = {
                'status': 'success',
                'message': 'Follow up details added'
            }
            return JsonResponse(response_data, status=200)

    except Exception as e:
        if type(e) is ValidationError:
            errormsg = '; '.join(ValidationError(e).messages)
        else:
            errormsg = str(e)

        response_data = {
            'status': 'fail',
            'message': "An error has occurred: " + errormsg
        }
        return JsonResponse(response_data)


def update_follow_up(request):
    try:
        if is_valid_post_request(request):
            follow_up_id = int(request.POST['follow_up_id'])
            follow_up = ClientFollowUp.objects.get(id=follow_up_id)

            if follow_up is not None:
                follow_up_type = ClientFollowUpType.objects.filter(id__exact=request.POST.get('follow_up_type')).first()
                follow_up_result_type = ClientLTFUResultType.objects.filter(id__exact=request.POST.get('follow_up_result_type')).first()
                follow_up_date = request.POST.get('follow_up_date')
                follow_up_comments = request.POST.get('follow_up_comments')

                follow_up.date_of_followup = follow_up_date
                follow_up.type_of_followup = follow_up_type
                follow_up.result_of_followup = follow_up_result_type
                follow_up.comment = follow_up_comments
                follow_up.save()

                response_data = {
                    'status': 'success',
                    'message': 'Follow up details updated'
                }
                return JsonResponse(response_data)
            else:
                response_data = {
                    'status': 'fail',
                    'message': "Error follow up not found"
                }
                return JsonResponse(response_data)
    except Exception as e:
        if type(e) is ValidationError:
            errormsg = '; '.join(ValidationError(e).messages)
        else:
            errormsg = str(e)

        response_data = {
            'status': 'fail',
            'message': "An error has occurred: " + errormsg
        }
        return JsonResponse(response_data)


def update_intervention(request):
    try:
        if request.method == 'POST' and request.user is not None and request.user.is_authenticated() and \
                request.user.is_active and request.user.has_perm('DreamsApp.change_intervention'):
            # Check if user belongs to an Ip
            if request.user.implementingpartneruser.implementing_partner is not None:
                intervention_id = int(request.POST.get('intervention_id'))
                if intervention_id is not None and type(intervention_id) is int:
                    intervention = Intervention.objects.get(id__exact=intervention_id)

                    if not intervention.is_editable_by_ip(request.user.implementingpartneruser.implementing_partner):
                        raise Exception(
                            'You do not have the rights to update this intervention.'
                        )

                    # check if intervention belongs to the ip
                    if intervention.implementing_partner == request.user.implementingpartneruser.implementing_partner:
                        intervention.intervention_type = InterventionType.objects.get(
                            code__exact=int(request.POST.get('intervention_type_code')))
                        intervention.client = Client.objects.get(id__exact=int(request.POST.get('client')))

                        intervention_date = dt.strptime(request.POST.get('intervention_date'), '%Y-%m-%d').date()

                        # check if external organisation is selected
                        external_organization_checkbox = request.POST.get('external_organization_checkbox')
                        external_organization_code = request.POST.get('external_organization_code')
                        other_external_organization_code = request.POST.get('other_external_organization_code')

                        if external_organization_checkbox:
                            if not external_organization_code:
                                response_data = {
                                    'status': 'fail',
                                    'message': "Error: External organisation must be selected if checkbox is checked."
                                }
                                return JsonResponse(response_data)
                        else:
                            if intervention.client.date_of_enrollment is not None and intervention_date < intervention.client.date_of_enrollment:
                                response_data = {
                                    'status': 'fail',
                                    'message': "Error: The intervention date must be after the client's enrollment date. "
                                }
                                return JsonResponse(response_data)

                        intervention.name_specified = request.POST.get('other_specify',
                                                                       '') if intervention.intervention_type.is_specified else ''
                        intervention.intervention_date = str(intervention_date)
                        intervention.changed_by = User.objects.get(id__exact=int(request.POST.get('changed_by')))
                        intervention.date_changed = dt.now()
                        intervention.comment = request.POST.get('comment')

                        i_type = InterventionType.objects.get(id__exact=intervention.intervention_type.id)

                        if i_type.has_hts_result:
                            intervention.hts_result = HTSResult.objects.get(
                                code__exact=int(request.POST.get('hts_result')))

                        if i_type.has_pregnancy_result:
                            intervention.pregnancy_test_result = PregnancyTestResult.objects.get(
                                code__exact=int(request.POST.get('pregnancy_test_result')))

                        if i_type.has_ccc_number:
                            intervention.client_ccc_number = request.POST.get('client_ccc_number')

                        if i_type.has_no_of_sessions:
                            intervention.no_of_sessions_attended = request.POST.get('no_of_sessions_attended')

                        if external_organization_checkbox:
                            intervention.external_organisation = ExternalOrganisation.objects.get(pk=external_organization_code)
                            if other_external_organization_code:
                                intervention.external_organisation_other = other_external_organization_code
                            else:
                                intervention.external_organisation_other = None

                        intervention.save(user_id=request.user.id, action="UPDATE")  # Logging
                        # using defer() miraculously solved serialization problem of datetime properties.
                        intervention = Intervention.objects.defer('date_changed', 'intervention_date',
                                                                  'date_created').get(id__exact=intervention.id)
                        # construct response

                        response_data = {
                            'status': 'success',
                            'message': 'Intervention successfully updated',
                            'intervention': serializers.serialize('json', [intervention, ], ensure_ascii=False),
                            'i_type': serializers.serialize('json', [i_type]),
                            'hts_results': serializers.serialize('json', HTSResult.objects.all()),
                            'pregnancy_results': serializers.serialize('json', PregnancyTestResult.objects.all()),
                            'permissions': json.dumps({
                                'can_change_intervention': request.user.has_perm('DreamsApp.change_intervention'),
                                'can_delete_intervention': request.user.has_perm('DreamsApp.delete_intervention')
                            })
                        }
                    else:
                        # Intervention does not belong to Implementing partner. Send back error message
                        raise Exception(
                            'You do not have the rights to update this intervention because it was created by a '
                            'different Implementing Partner'
                        )
                    return JsonResponse(response_data)
                else:
                    raise Exception("Error: No intervention selected for update")
            else:  # Raise an exception. User does not belong to any IP
                raise Exception("Error: You do not belong to an Implementing Partner. "
                                "Please contact your system admin to add you to the relevant Implementing Partner.")
        else:
            response_data = {
                'status': 'fail',
                'message': "You Cannot edit interventions. Please contact System Administrator for help."
            }
            return JsonResponse(response_data)
    except Exception as e:
        # check if validation error
        if type(e) is ValidationError:
            errormsg = '; '.join(ValidationError(e).messages)
        else:
            errormsg = str(e)

        response_data = {
            'status': 'fail',
            'message': "An error has occurred: " + errormsg
        }
        return JsonResponse(response_data)


def delete_follow_up(request):
    try:
        if is_valid_post_request(request):
            follow_up_id = int(request.POST.get('follow_up_id'))

            if follow_up_id is not None and type(follow_up_id) is int:
                follow_up = ClientFollowUp.objects.filter(pk=follow_up_id).first()

                if follow_up is not None:
                    ClientFollowUp.objects.filter(pk=follow_up_id).delete()
                    log_custom_actions(request.user.id, "DreamsApp_clientfollowup", follow_up_id, "DELETE", None)

                    response_data = {
                        'status': 'success',
                        'message': 'Follow up has been successfully deleted',
                        'follow_up_id': follow_up_id
                    }
                    return JsonResponse(response_data)
                else:
                    response_data = {
                        'status': 'fail',
                        'message': 'Follow up not found'
                    }
                    return JsonResponse(response_data)

    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. "
                       "Please contact the System Administrator if this error persists."
        }
        return JsonResponse(response_data)


def delete_intervention(request):
    try:
        if request.method == 'POST' and request.user is not None and request.user.is_authenticated() and \
                request.user.is_active and request.user.has_perm('DreamsApp.delete_intervention'):
            # Check if user belongs to an Ip
            if request.user.implementingpartneruser.implementing_partner is not None:
                intervention_id = int(request.POST.get('intervention_delete_id'))
                if intervention_id is not None and type(intervention_id) is int:
                    # get intervention
                    # Check if intervention belongs to IP
                    intervention = Intervention.objects.filter(pk=intervention_id).first()

                    if not intervention.is_editable_by_ip(request.user.implementingpartneruser.implementing_partner):
                        response_data = {
                            'status': 'fail',
                            'message': 'You do not have the rights to delete this intervention.'
                        }
                        return JsonResponse(response_data)

                    if intervention.implementing_partner == request.user.implementingpartneruser.implementing_partner:
                        intervention.voided = True
                        intervention.voided_by = request.user
                        intervention.date_voided = datetime.now()
                        intervention.save(user_id=request.user.id, action="UPDATE")  # Updating logs
                        # intervention.delete() # No deletion whatsoever
                        log_custom_actions(request.user.id, "DreamsApp_intervention", intervention_id, "DELETE", None)
                        response_data = {
                            'status': 'success',
                            'message': 'Intervention has been successfully deleted',
                            'intervention_id': intervention_id
                        }
                        return JsonResponse(response_data)
                    else:
                        response_data = {
                            'status': 'fail',
                            'message': 'You do not have the rights to delete this intervention because it was created by a '
                                       'different Implementing Partner'
                        }
                        return JsonResponse(response_data)
                else:
                    response_data = {
                        'status': 'fail',
                        'message': "Error: No intervention selected for deletion"
                    }
                    return JsonResponse(response_data)
            else:
                response_data = {
                    'status': 'fail',
                    'message': "Error: You do not belong to an Implementing Partner. "
                               "Please contact your system admin to add you to the relevant Implementing Partner."
                }
                return JsonResponse(response_data)
        else:
            response_data = {
                'status': 'fail',
                'message': "You don't have permission to delete Intervention. Please contact System Administrator for help."
            }
            return JsonResponse(response_data)
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. "
                       "Contact System Administrator if this error Persists."
        }
        return JsonResponse(response_data)


def get_sub_counties(request):
    try:
        if request.method == 'GET' and request.user is not None and request.user.is_authenticated() and request.user.is_active:
            response_data = {}
            county_id = request.GET['county_id']
            county = County.objects.get(id__exact=county_id)
            sub_counties = SubCounty.objects.filter(county__exact=county.id)
            sub_counties = serializers.serialize('json', sub_counties)
            response_data["sub_counties"] = sub_counties
            return JsonResponse(response_data)
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)


def get_wards(request):
    if request.method == 'GET' and request.user is not None and request.user.is_authenticated() and request.user.is_active:
        response_data = {}
        sub_county_id = request.GET['sub_county_id']
        sub_county = SubCounty.objects.get(id__exact=sub_county_id)
        wards = Ward.objects.filter(sub_county__exact=sub_county.id)
        wards = serializers.serialize('json', wards)
        response_data["wards"] = wards
        return JsonResponse(response_data)

    else:
        raise PermissionDenied


def log_me_out(request):
    if request.user.is_authenticated():
        logout(request)
        return redirect('login')
    else:
        raise PermissionDenied


def reporting(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if request.method == 'GET':
                return render(request, 'reporting.html', {'user': request.user, 'page_title': 'DREAMS Reporting', })
            elif request.method == 'POST' and request.is_ajax():
                return render(request, 'reporting.html', {'user': request.user, 'page_title': 'DREAMS Reporting', })
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc()
        return HttpResponseServerError(tb)  # for debugging purposes. Will only report exception


def user_help(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if request.method == 'GET':
                return render(request, 'help.html', {
                    'user': request.user,
                    'page': 'help',
                    'page_title': 'DREAMS User help'
                })
            elif request.method == 'POST' and request.is_ajax():
                return render(request, 'help.html', {
                    'user': request.user,
                    'page': 'help',
                    'page_title': 'DREAMS User help'
                })
        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)  # for debugging purposes. Will only report exception


def user_help_download(request):
    if request.user.is_authenticated() and request.user.is_active:
        try:
            manual_filename = request.POST.get('manual') if request.method == 'POST' else request.GET.get('manual')
            manual_friendly_name = request.POST.get(
                'manual_friendly_name') if request.method == 'POST' else request.GET.get('manual_friendly_name')
            fs = FileSystemStorage(location=os.path.join(settings.BASE_DIR, 'templates', 'manuals'))
            com_path = fs.location
            filename = manual_filename + '.pdf'
            if fs.exists(filename):
                with fs.open(filename) as pdf:
                    response = HttpResponse(pdf, content_type='application/pdf')
                    response['Content-Disposition'] = 'attachment; filename="' + manual_friendly_name + '"'
                    return response
            else:
                traceback.format_exc()
            return
        except Exception as e:
            traceback.format_exc()
            return
    else:
        raise PermissionDenied


def logs(request):
    if request.user.is_authenticated() and request.user.is_active:
        if not request.user.is_superuser and not request.user.has_perm('DreamsApp.can_manage_audit'):
            raise PermissionDenied('Operation not allowed. [Missing Permission]')

        ip = ''
        try:
            ip = request.user.implementingpartneruser.implementing_partner.id
        except ImplementingPartnerUser.DoesNotExist:
            pass

        # user is allowed to view logs
        if request.method == 'GET':
            try:
                page = request.GET.get('page', 1)
                filter_text = request.GET.get('filter-log-text', '')
                filter_date_from = request.GET.get('filter-log-date-from', '')
                filter_date = request.GET.get('filter-log-date', '')

                # getting logs
                logs = Audit.objects.filter(Q(table__in=filter_text.split(" ")) |
                                            Q(action__in=filter_text.split(" ")) |
                                            Q(search_text__in=filter_text.split(" ")) |
                                            Q(user__username__icontains=filter_text.split(" ")[0]) |
                                            Q(user__first_name__icontains=filter_text.split(" ")[0]) |
                                            Q(user__last_name__icontains=filter_text.split(" ")[0])
                                            ).order_by('-timestamp')

                logs = filter_audit_logs_by_date_and_ip(filter_date, filter_date_from, ip, logs)

                paginator = Paginator(logs, 25)  # Showing 25 contacts per page
                try:
                    logs_list = paginator.page(page)
                except PageNotAnInteger:
                    logs_list = paginator.page(1)  # Deliver the first page is page is not an integer
                except EmptyPage:
                    logs_list = paginator.page(0)  # Deliver the last page if page is out of scope
                return render(request, 'log.html', {'page': 'logs', 'page_title': 'DREAMS Logs', 'logs': logs_list,
                                                    'filter_text': filter_text,
                                                    'filter_date_from': filter_date_from,
                                                    'filter_date': filter_date,
                                                    'items_in_page': 0 if logs_list.end_index() == 0 else
                                                    (logs_list.end_index() - logs_list.start_index() + 1)
                                                    }
                              )
            except Exception as e:
                tb = traceback.format_exc(e)
                return HttpResponseServerError(tb)
        elif request.method == 'POST':
            # get the form data
            filter_text = request.POST.get('filter-log-text', '')
            filter_date = request.POST.get('filter-log-date', '')
            filter_date_from = request.POST.get('filter-log-date-from', '')

            logs = Audit.objects.filter(Q(table__icontains=filter_text) |
                                        Q(action__icontains=filter_text) |
                                        Q(search_text__icontains=filter_text) |
                                        Q(user__username__icontains=filter_text) |
                                        Q(user__first_name__icontains=filter_text) |
                                        Q(user__last_name__icontains=filter_text)).order_by('-timestamp')

            logs = filter_audit_logs_by_date_and_ip(filter_date, filter_date_from, ip, logs)

            paginator = Paginator(logs, 25)
            try:
                logs_list = paginator.page(1)
            except PageNotAnInteger:
                logs_list = paginator.page(1)  # Deliver the first page is page is not an integer
            except EmptyPage:
                logs_list = paginator.page(0)  # Deliver the last page if page is out of scope
            return render(request, 'log.html', {'page': 'logs',
                                                'page_title': 'DREAMS Logs',
                                                'logs': logs_list,
                                                'filter_text': filter_text,
                                                'filter_date_from': filter_date_from,
                                                'filter_date': filter_date,
                                                'items_in_page': 0 if logs_list.end_index() == 0 else
                                                (logs_list.end_index() - logs_list.start_index() + 1)})
    else:
        raise PermissionDenied


def filter_audit_logs_by_date_and_ip(filter_date, filter_date_from, ip, logs):
    if ip != '':
        logs = logs.filter(Q(user__implementingpartneruser__implementing_partner__id__exact=ip))
    if filter_date_from == '' and filter_date == '':
        pass
    elif filter_date_from != '' and filter_date == '':
        fyr, fmnth, fdt = filter_date_from.split('-')
        constructed_date_from = date(int(fyr), int(fmnth), int(fdt))
        logs = logs.filter(Q(timestamp__date__gte=constructed_date_from))
    elif filter_date_from == '' and filter_date != '':
        yr, mnth, dat = filter_date.split('-')
        constructed_date = date(int(yr), int(mnth), int(dat))
        logs = logs.filter(Q(timestamp__date__lte=constructed_date))
    else:
        yr, mnth, dat = filter_date.split('-')
        constructed_date = date(int(yr), int(mnth), int(dat))
        fyr, fmnth, fdt = filter_date_from.split('-')
        constructed_date_from = date(int(fyr), int(fmnth), int(fdt))
        logs = logs.filter(Q(timestamp__date__gte=constructed_date_from) &
                           Q(timestamp__date__lte=constructed_date))
    return logs


# user management
def users(request):
    if request.user.is_authenticated() and request.user.is_active and request.user.has_perm('auth.can_manage_user'):
        try:
            items_per_page = 15
            if request.method == 'GET':
                page = request.GET.get('page', 1)
                filter_text = request.GET.get('filter-user-text', '')
                ip_user_list = ImplementingPartnerUser.objects.filter(Q(user__first_name__contains=filter_text) |
                                                                      Q(user__last_name__contains=filter_text) |
                                                                      Q(user__username__contains=filter_text)
                                                                      ).order_by('-user__date_joined')
            elif request.method == 'POST':
                page = request.POST.get('page', 1)
                filter_text = request.POST.get('filter-user-text', '')
                # get list of users with the filter incorporated
                if filter_text is not u'' and filter_text is not " ":
                    ip_user_list = ImplementingPartnerUser.objects.filter(
                        Q(user__first_name__in=filter_text.split(" ")) |
                        Q(user__last_name__in=filter_text.split(" ")) |
                        Q(user__username__in=filter_text.split(" "))
                    ).order_by('-user__date_joined')
                else:
                    ip_user_list = ImplementingPartnerUser.objects.filter(Q(user__first_name__contains=filter_text) |
                                                                          Q(user__last_name__contains=filter_text) |
                                                                          Q(user__username__contains=filter_text)
                                                                          ).order_by('-user__date_joined')
            #  current user ip
            try:
                current_user_ip = request.user.implementingpartneruser.implementing_partner
            except ImplementingPartnerUser.DoesNotExist:
                current_user_ip = None
            except ImplementingPartner.DoesNotExist:
                current_user_ip = None
            if not request.user.has_perm('DreamsApp.can_view_cross_ip_data'):
                if current_user_ip is not None:
                    ip_user_list = ip_user_list.filter(implementing_partner=current_user_ip)
                else:
                    ip_user_list = ImplementingPartnerUser.objects.none()
            # do pagination
            paginator = Paginator(ip_user_list, items_per_page)
            final_ip_user_list = paginator.page(page)
        except ImplementingPartnerUser.DoesNotExist:
            current_user_ip = None
        except PageNotAnInteger:
            final_ip_user_list = paginator.page(1)  # Deliver the first page is page is not an integer
        except EmptyPage:
            final_ip_user_list = paginator.page(0)  # Deliver the last page if page is out of scope
        return render(request, 'users.html',
                      {'page': 'users', 'page_title': 'DREAMS User List', 'ip_users': final_ip_user_list,
                       'filter_text': filter_text,
                       'items_in_page': 0 if final_ip_user_list.end_index() == 0 else
                       (final_ip_user_list.end_index() - final_ip_user_list.start_index() + 1),
                       'implementing_partners': ImplementingPartner.objects.all(),
                       'current_user_ip': current_user_ip,
                       'roles': Group.objects.all()
                       })
    else:
        raise PermissionDenied  # this should be a redirection to the permissions denied page


def save_user(request):
    if request.method == 'POST' and request.user is not None and request.user.is_authenticated() and request.user.is_active and request.user.has_perm(
            'auth.can_manage_user'):
        try:
            new_user_ip = ImplementingPartner.objects.get(name=
                                                          request.POST.get('implementing_partner',
                                                                           ''))  # Valid IP for new user
            # check if user can change cross IP data
            if not request.user.has_perm('auth.can_change_cross_ip_data'):
                # User must register new user under their IP. Theck if user has a valid IP
                # check if registering user belongs to an IP
                if request.user.implementingpartneruser.implementing_partner is None:  # Registering user does not belong to an IP. Raise exception
                    raise Exception("Error: You do not belong to an Implementing Partner. "
                                    "Please contact your system admin to add you to the relevant Implementing Partner.")

                elif new_user_ip != request.user.implementingpartneruser.implementing_partner:
                    # Registering user and user do not belong to same IP. Raise exception
                    raise Exception(
                        "Error: You do not have permission to register a user under a different Implementing"
                        " partner other than " + request.user.implementingpartneruser.
                        implementing_partner.name)

            # Everything is fine. Proceed to register user
            # Check to see if IP user exists already or not
            ip_user_id = request.POST.get('ip_user_id', 0)
            if ip_user_id == 0 or ip_user_id is None or ip_user_id == u'':
                # This is a new user
                username = request.POST.get('username', '')
                email = request.POST.get('emailaddress', '')
                role = request.POST.get('role', '')
                # Check if provided username is in db
                try:
                    user = User.objects.get(username__exact=username)
                    raise Exception("Error: The username '" + username +
                                    "' is already take. Please enter a different username and continue")
                except User.DoesNotExist:
                    # user with this username does not exist
                    user_group = Group.objects.filter(name__exact=role).first()  # get group
                    user = User.objects.create_user(
                        username=username, email=email, password='PTZ%hz^+&mny+9Av'
                    )
                    user.first_name = request.POST.get('firstname', '')
                    user.last_name = request.POST.get('lastname', '')
                    user.groups.add(user_group)
                    user.save()  # save user changes first
                    ip_user = ImplementingPartnerUser.objects.create(
                        user=user,
                        implementing_partner=new_user_ip
                    )
                    try:
                        # send email with user credentials
                        msg = EmailMessage('DREAMS Credentials',
                                           'Dear ' + user.first_name + ',\n\n Welcome to DREAMS. You can login to your account at http://dreams-dev.globalhealthapp.net.'
                                                                       '\n\n\t Username: ' + username +
                                           '\n\t Password:  PTZ%hz^+&mny+9Av'
                                           '\n\nThank you,'
                                           '\nDREAMS Administrator',
                                           to=[email])
                        msg.send()
                        response_data = {
                            'status': 'success',
                            'message': 'User successfully saved. Your temporary password has been sent to ' + email,
                            'ip_users': serializers.serialize('json', [ip_user, ])
                        }
                        return JsonResponse(response_data)
                    except Exception as e:
                        response_data = {
                            'status': 'success',
                            'message': 'User registered but could not send login detais to provided email address. '
                                       'Contact System Admin for help ',
                            'ip_users': serializers.serialize('json', [ip_user, ])
                        }
                        return JsonResponse(response_data)
            else:  # raise exception
                raise Exception('User with id ' + ip_user_id + ' exists already')
        except ImplementingPartner.DoesNotExist:
            # User being registered under invalid IP. Raise an error
            raise Exception("Error: User must be registered under an Implementing Partner")

        except Exception as e:
            response_data = {
                'status': 'fail',
                'message': "An error occurred while processing request. Contact System Administrator if this error Persists.",
                'ip_users': ''
            }
            return JsonResponse(response_data)
    else:
        raise PermissionDenied


def toggle_status(request):
    try:
        if request.method == 'GET' and request.user is not None and request.user.is_authenticated() \
                and request.user.is_active and request.user.has_perm('auth.can_change_user_status'):
            if request.method == 'GET':
                ip_user_id = request.GET.get('ip_user_id', 0)
                toggle = str(request.GET.get('toggle', False))
                #  get ip user
                ip_user = ImplementingPartnerUser.objects.filter(id__exact=ip_user_id).first()
                if ip_user is not None:
                    # Get user object
                    user = User.objects.filter(id__exact=ip_user.user.id).first()
                    if user is not None:
                        # check if user is same as requesting user
                        if user.id == request.user.id:
                            raise Exception('Error: you cannot deactivate your own account. '
                                            'Please contact System Administrator for assistance')
                        else:
                            user.is_active = toggle in ["True", "true", 1, "Yes", "yes", "Y", "y", "T", "t"]
                            user.save()
                            # return success message
                            response_data = {
                                'status': 'success',
                                'message': 'User status changed successfully'
                            }
                            return JsonResponse(response_data)
                    else:
                        raise Exception('Error: you did not select a valid user. Please try again')
                else:
                    raise Exception('Error: you did not select a valid user. Please try again')
            else:
                raise Exception("Invalid user request!")
        else:
            raise Exception('Error: You do not have permission to perform this operation. Please contact your server '
                            'administrator for assistance')
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. Contact System Administrator if this error Persists."
        }
        return JsonResponse(response_data)


def change_cred(request):
    if request.user.is_authenticated() and request.user.is_active:  # user is authenticated
        if request.method == 'GET':
            context = {'page': 'account', 'page_title': 'DREAMS Password Change', 'user': request.user, }
            return render(request, 'change_cred.html', context)
        elif request.method == 'POST':
            ch_username = request.POST.get('ch_username', '')
            ch_current_password = request.POST.get('ch_current_password', '')
            ch_new_password = request.POST.get('ch_new_password', '')
            ch_confirm_new_password = request.POST.get('ch_confirm_new_password', '')
            if ch_username == '' or ch_current_password == '' or ch_new_password == '' or ch_confirm_new_password == '' or ch_confirm_new_password == '':
                response_data = {
                    'status': 'fail',
                    'message': 'An error occurred as a result of missing details.'
                }
                return JsonResponse(response_data)
            else:
                if request.user.get_username() == ch_username and request.user.check_password(
                        ch_current_password) and ch_new_password == ch_confirm_new_password:
                    request.user.set_password(ch_new_password)
                    request.user.save()
                    response_data = {
                        'status': 'success',
                        'message': 'Password changed successfully. Proceed to login with your new credentials'
                    }
                    return JsonResponse(response_data)

                else:
                    response_data = {
                        'status': 'fail',
                        'message': 'An error occurred as a result of wrong credentials.'
                    }
                    return JsonResponse(response_data)

    else:  # User is not logged in. Redirect user to login page
        # return PermissionDenied('Operation not allowed. [Missing Permission]')
        raise PermissionDenied


def bad_request(request):
    context = {'user': request.user, 'error_code': 400, 'error_title': 'Bad Request (Error 400)',
               'error_message':
                   'Your browser sent a request that this server could not understand<br>. '}
    return render(request, 'error_page.html', context)


def permission_denied(request):  # PermissionDenied('Operation not allowed. [Missing Permission]')
    context = {'user': request.user, 'error_code': 403, 'error_title': 'Permission Denied (Error 403)',
               'error_message':
                   'You do not have permission to view this page [Missing Permission]. '
                   'Go back to previous page or Home page'}
    return render(request, 'error_page.html', context)


def page_not_found(request):
    context = {'user': request.user, 'error_code': 404, 'error_title': 'Page Not Found (Error 404)',
               'error_message': 'The page you are looking for does not exist. Go back to previous page or Home page'}
    return render(request, 'error_page.html', context)


def server_error(request):
    context = {'user': request.user, 'error_code': 500, 'error_title': 'Server Error (Error 500)',
               'error_message':
                   'A server error occurred while processing your request. Please try again or contact your '
                   'administrator if the error persists.<br>. '}
    return render(request, 'error_page.html', context)


def grievances_list(request):
    try:
        if not request.user.is_authenticated():
            raise PermissionDenied
        page = request.GET.get('page', 1) if request.method == 'GET' else request.POST.get('page', 1)
        filter_date = request.GET.get('filter_date', None) if request.method == 'GET' else request.POST.get(
            'filter_date', None)
        filter_text = request.GET.get('filter-user-text', '') if request.method == 'GET' else request.POST.get(
            'filter-user-text', '')
        try:
            user_ip = request.user.implementingpartneruser.implementing_partner
        except:
            user_ip = None
        """ IP level permission check """
        grievance_list = Grievance.objects.all() if request.user.has_perm('DreamsApp.can_view_cross_ip_data') else \
            Grievance.objects.filter(implementing_partner=user_ip)
        """Date filter """
        if filter_date is not None and filter_date is not u'':
            yr, mnth, dt = filter_date.split('-')
            constructed_date = date(int(yr), int(mnth), int(dt))
            grievance_list = Grievance.objects.filter(Q(date__year=constructed_date.year,
                                                        date__month=constructed_date.month,
                                                        date__day=constructed_date.day))
        """ Text filter """
        # if filter_text is not u'' and filter_text is not '':
        # grievance_list = Grievance.objects.filter(Q(reporter_name__contains=filter_text) |
        #                                         Q(relationship__containts=filter_text) |
        #                                        Q(reporter_phone__contains=filter_text))
        # do pagination
        try:
            paginator = Paginator(grievance_list, 20)
            final_grievance_list = paginator.page(page)
        except PageNotAnInteger:
            final_grievance_list = paginator.page(1)  # Deliver the first page is page is not an integer
        except EmptyPage:
            final_grievance_list = paginator.page(0)  # Deliver the last page if page is out of scope
        response_data = {
            'page': 'cash_transfer', 'page_title': 'DREAMS Grievance List',
            'filter_text': filter_text,
            'filter_date': filter_date,
            'items_in_page': 0 if final_grievance_list.end_index() == 0 else
            (final_grievance_list.end_index() - final_grievance_list.start_index() + 1),
            'current_user_ip': user_ip,
            'form': GrievanceModelForm(current_user=request.user),
            'grievance_list': final_grievance_list,
            'status': 'success'
        }
        return render(request, 'grievances.html', response_data)
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. Contact System Administrator if this error Persists."
        }
        return JsonResponse(response_data)


@csrf_exempt
def grievances_create(request):
    try:
        if request.user.is_authenticated() and request.method == 'POST' and request.is_ajax():
            grievance = GrievanceModelForm(request.POST, current_user=request.user)
            if grievance.is_valid():
                saved_grievance = grievance.save(commit=False)
                saved_grievance.created_by = request.user
                saved_grievance.save()
                response_data = {
                    'status': 'success',
                    'message': 'Grievance saved successfully',
                    'grievance': grievance.data,
                    'grievance_id': saved_grievance.id,
                    'reporter_categories': serializers.serialize('json', GrievanceReporterCategory.objects.all()),
                    'grievance_nature': serializers.serialize('json', GrievanceNature.objects.all()),
                    'status_list': serializers.serialize('json', GrievanceStatus.objects.all()),
                }
                return JsonResponse(response_data)
            else:
                raise Exception(grievance.errors)
        else:
            raise PermissionDenied
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. Contact System Administrator if this error Persists."
        }
        return JsonResponse(response_data)


@csrf_exempt
def grievances_edit(request):
    """ """
    try:
        if request.user.is_authenticated() and request.method == 'POST' and request.is_ajax():
            try:
                id = int(request.POST.get('id', 0))
                instance = Grievance.objects.get(id=id)
                form = GrievanceModelForm(request.POST, instance=instance, current_user=request.user)
                if not form.is_valid():
                    response_data = {
                        'status': 'fail',
                        'message': form.errors
                    }
                else:
                    saved_grievance = form.save(commit=False)
                    saved_grievance.changed_by = request.user
                    saved_grievance.save()
                    response_data = {
                        'status': 'success',
                        'message': 'Grievance edited successfully',
                        'grievance': form.data,
                        'reporter_categories': serializers.serialize('json', GrievanceReporterCategory.objects.all()),
                        'grievance_nature': serializers.serialize('json', GrievanceNature.objects.all()),
                        'status_list': serializers.serialize('json', GrievanceStatus.objects.all()),
                    }
                    return JsonResponse(response_data)
            except Grievance.DoesNotExist:
                response_data = {
                    'status': 'fail',
                    'message': 'Grievance not found'
                }
                return JsonResponse(response_data)
        else:
            raise PermissionDenied
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. Contact System Administrator if this error Persists."
        }
        return JsonResponse(response_data)


@csrf_exempt
def grievances_delete(request):
    """ """
    try:
        if request.user.is_authenticated() and request.method == 'POST' and request.is_ajax():
            try:
                id = int(request.POST.get('id', 0))
                instance = Grievance.objects.get(id=id).delete()
                response_data = {
                    'status': 'success',
                    'message': 'Grievance deleted successfully',
                    'grievanceId': id
                }
                return JsonResponse(response_data)
            except Grievance.DoesNotExist:
                response_data = {
                    'status': 'fail',
                    'message': 'Grievance not found'
                }
                return JsonResponse(response_data)
        else:
            raise PermissionDenied
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. Contact System Administrator if this error Persists."
        }
        return JsonResponse(response_data)


def grievances_get(request):
    """ """
    try:
        if not request.user.is_authenticated():
            raise PermissionDenied
        grievance_id = request.GET.get('grievance_id', 0) if request.method == 'GET' else request.POST.get(
            'grievance_id', 0)
        try:
            grievance = Grievance.objects.get(id__exact=grievance_id)
            response_data = {
                'status': 'success',
                'message': 'Grievance saved successfully',
                'grievance': serializers.serialize('json', [grievance]),
            }
        except Grievance.DoesNotExist:
            response_data = {
                'status': 'fail',
                'message': 'Grievance not found in Database'
            }
        return JsonResponse(response_data)
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. Contact System Administrator if this error Persists."
        }
        return JsonResponse(response_data)


@csrf_exempt
def cash_transfer_details_save(request):
    try:
        if request.user.is_authenticated() and request.method == 'POST' and request.is_ajax():
            # check if details has id
            id = request.POST.get('id', 0)
            if id not in [u'', "0", 0, ""]:
                id = int(id)
                ct_detail = ClientCashTransferDetails.objects.get(id=id)
                ct_form = ClientCashTransferDetailsForm(request.POST, instance=ct_detail)
            else:
                # get current AGYW
                ct_form = ClientCashTransferDetailsForm(request.POST)
            if ct_form.is_valid():
                saved_detail = ct_form.save(commit=False)
                if saved_detail.id is not None and saved_detail.id > 0:
                    saved_detail.changed_by = request.user
                else:
                    saved_detail.created_by = request.user
                saved_detail.save()
                response_data = {
                    'status': 'success',
                    'message': 'Cash Transfer details updated successfully',
                    'ct_detail_id': saved_detail.id,
                }
                return JsonResponse(response_data)
            else:
                raise Exception(ct_form.errors)
        else:
            raise PermissionDenied
    except ClientCashTransferDetails.DoesNotExist:
        raise Exception("Cash transfer details does not exist for editing!")
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': "An error occurred while processing request. Contact System Administrator if this error Persists."
        }
        return JsonResponse(response_data)


def export_page(request):
    if request.user.is_authenticated() and request.user.is_active and request.user.has_perm(
            'DreamsApp.can_export_raw_data'):

        try:
            ips = None
            if request.user.is_superuser or request.user.has_perm('DreamsApp.can_view_cross_ip_data'):
                ips = ImplementingPartner.objects.all()

            elif request.user.implementingpartneruser is not None:
                ips = ImplementingPartner.objects.filter(
                    id=request.user.implementingpartneruser.implementing_partner.id)

                if ips.count() > 0:
                    sub_grantees = ImplementingPartner.objects.filter(parent_implementing_partner__in=ips)
                    if sub_grantees.exists():
                        ips = ips.union(sub_grantees)

            context = {'page': 'export', 'page_title': 'Client Raw Enrolment Export', 'ips': ips,
                       'counties': County.objects.all()}

            return render(request, 'dataExport.html', context)
        except ImplementingPartnerUser.DoesNotExist:
            traceback.format_exc()
        except ImplementingPartner.DoesNotExist:
            traceback.format_exc()
    else:
        raise PermissionDenied


def intervention_export_page(request):
    if request.user.is_authenticated() and request.user.is_active and request.user.has_perm(
            'DreamsApp.can_export_raw_data'):

        try:
            ips = None
            if request.user.is_superuser or request.user.has_perm('DreamsApp.can_view_cross_ip_data'):
                ips = ImplementingPartner.objects.all()
            elif request.user.implementingpartneruser is not None:
                ips = ImplementingPartner.objects.filter(
                    id=request.user.implementingpartneruser.implementing_partner.id)

                if ips.count() > 0:
                    sub_grantees = ImplementingPartner.objects.filter(parent_implementing_partner__in=ips)
                    if sub_grantees.exists():
                        ips = ips.union(sub_grantees)

            context = {'page': 'export', 'page_title': 'Client Interventions Export', 'ips': ips,
                       'counties': County.objects.all()}
            return render(request, 'interventionDataExport.html', context)
        except ImplementingPartnerUser.DoesNotExist:
            traceback.format_exc()
        except ImplementingPartner.DoesNotExist:
            traceback.format_exc()
    else:
        raise PermissionDenied


def download_raw_enrollment_export(request):
    try:
        ip_list_str = request.POST.getlist('ips')
        sub_county = request.POST.get('sub_county')
        ward = request.POST.get('ward')
        county = request.POST.get('county_of_residence')
        export_file_name = urllib.parse.quote(("/tmp/raw_enrolment_export-{}.csv").format(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        export_doc = DreamsRawExportTemplateRenderer()

        if request.user.is_superuser or request.user.has_perm('DreamsApp.can_view_phi_data') \
                or Permission.objects.filter(group__user=request.user).filter(
            codename='DreamsApp.can_view_phi_data').exists():
            show_PHI = True
        else:
            show_PHI = False

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = ('attachment; filename="{}"').format(export_file_name)
        export_doc.prepare_enrolment_export_doc(response, ip_list_str, sub_county, ward, show_PHI)

        return response

    except Exception as e:
        traceback.format_exc()
        return


def download_raw_intervention_export(request):
    try:
        ip_list_str = request.POST.getlist('ips')
        sub_county = request.POST.get('sub_county')
        ward = request.POST.get('ward')
        county = request.POST.get('county_of_residence')
        export_file_name = urllib.parse.quote(
            ("/tmp/raw_intervention_export-{}.csv").format(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        export_doc = DreamsRawExportTemplateRenderer()

        if request.user.is_superuser or request.user.has_perm('DreamsApp.can_view_phi_data') \
                or Permission.objects.filter(group__user=request.user).filter(
            codename='DreamsApp.can_view_phi_data').exists():
            show_PHI = True
        else:
            show_PHI = False

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = ('attachment; filename="{}"').format(export_file_name)
        export_doc.get_intervention_excel_doc(response, ip_list_str, sub_county, ward, show_PHI)

        return response
    except Exception as e:
        traceback.format_exc()
        return


def individual_service_layering_export_page(request):
    if request.user.is_authenticated() and request.user.is_active and request.user.has_perm(
            'DreamsApp.can_export_raw_data'):

        try:
            ips = None
            if request.user.is_superuser or request.user.has_perm('DreamsApp.can_view_cross_ip_data'):
                ips = ImplementingPartner.objects.all()
            elif request.user.implementingpartneruser is not None:
                ips = ImplementingPartner.objects.filter(
                    id=request.user.implementingpartneruser.implementing_partner.id)

                if ips.count() > 0:
                    sub_grantees = ImplementingPartner.objects.filter(parent_implementing_partner__in=ips)
                    if sub_grantees.exists():
                        ips = ips.union(sub_grantees)

            context = {'page': 'export', 'page_title': 'Service Layering Export', 'ips': ips,
                       'counties': County.objects.all()}
            return render(request, 'individualServiceLayeringDataExport.html', context)

        except ImplementingPartnerUser.DoesNotExist:
            traceback.format_exc()
        except ImplementingPartner.DoesNotExist:
            traceback.format_exc()
    else:
        raise PermissionDenied


def download_services_received_export(request):
    try:
        ip_list_str = request.POST.getlist('ips')
        sub_county = request.POST.get('sub_county')
        ward = request.POST.get('ward')
        county = request.POST.get('county_of_residence')
        export_file_name = urllib.parse.quote(
            ("/tmp/services_received_export-{}.csv").format(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        export_doc = DreamsRawExportTemplateRenderer()

        if request.user.is_superuser or request.user.has_perm('DreamsApp.can_view_phi_data') \
                or Permission.objects.filter(group__user=request.user).filter(
            codename='DreamsApp.can_view_phi_data').exists():
            show_PHI = True
        else:
            show_PHI = False

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = ('attachment; filename="{}"').format(export_file_name)
        export_doc.get_individual_layering_report(response, ip_list_str, sub_county, ward, show_PHI)
        return response

    except Exception as e:
        traceback.format_exc()
        return


def error_404(request):
    context = {'user': request.user, 'error_code': 404, 'page_title': 'DREAMS Application Error',
               'error_title': 'Page Not Found (Error 404)',
               'error_message': 'The page you are looking for does not exist. Go back to previous page or Home page'}
    return render(request, 'error_page.html', context)


def viewBaselineData(request):
    """ Returns client profile """
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:
        if request.method == 'GET':
            try:
                client_id = int(request.GET['client_id'])
                client_demographics = Client.objects.filter(id=client_id).exclude(voided=True).first()
                client_household = ClientIndividualAndHouseholdData.objects.filter(client=client_id).exclude(
                    voided=True).first()
                client_edu = ClientEducationAndEmploymentData.objects.filter(client=client_id).exclude(
                    voided=True).first()
                client_sexual_data = ClientSexualActivityData.objects.filter(client=client_id).exclude(
                    voided=True).first()
                client_gbv_data = ClientGenderBasedViolenceData.objects.filter(client=client_id).exclude(
                    voided=True).first()
                client_hiv_data = ClientHIVTestingData.objects.filter(client=client_id).exclude(voided=True).first()
                client_rh_data = ClientReproductiveHealthData.objects.filter(client=client_id).exclude(
                    voided=True).first()
                client_drug_data = ClientDrugUseData.objects.filter(client=client_id).exclude(voided=True).first()
                client_prog_part_data = ClientParticipationInDreams.objects.filter(client=client_id).exclude(
                    voided=True).first()

                demographics_form = DemographicsForm(instance=client_demographics)
                household_form = IndividualAndHouseholdForm(instance=client_household)
                edu_and_emp_form = EducationAndEmploymentForm(instance=client_edu)
                sexuality_form = SexualityForm(instance=client_sexual_data)
                gbv_form = GBVForm(instance=client_gbv_data)
                hiv_form = HivTestForm(instance=client_hiv_data)
                reproductive_health_form = ReproductiveHealthForm(instance=client_rh_data)
                drug_use_form = DrugUseForm(instance=client_drug_data)
                participation_form = DreamsProgramParticipationForm(instance=client_prog_part_data)

                search_client_term = request.GET.get('search_client_term', '')
            except Client.DoesNotExist:
                traceback.format_exc()
            except Exception as e:
                traceback.format_exc()
                print(str(e))
        else:
            print ('POST not allowed')

        ip = None
        try:
            ip = request.user.implementingpartneruser.implementing_partner
            if ip:
                ip_code = ip.code
            else:
                ip = None
        except Exception as e:
            ip_code = None

        if client_id is not None and client_id != 0:
            try:
                client_found = Client.objects.get(id=client_id)
                if client_found is not None:
                    is_editable_by_ip = client_found.is_editable_by_ip(ip)
                    client_status = client_found.get_client_status(ip)
                    date_of_enrollment_str = demographics_form['date_of_enrollment'].value()
                    date_of_enrollment = datetime.strptime(str(date_of_enrollment_str), '%Y-%m-%d').date() if date_of_enrollment_str is not None else dt.now().date()
                    client_enrolment_service_layer = ClientEnrolmentServiceLayer(request.user)
                    minimum_maximum_age = client_enrolment_service_layer.get_minimum_maximum_enrolment_age(
                        client_enrolment_service_layer.ENROLMENT_CUTOFF_DATE)
                    max_dob = date_of_enrollment - relativedelta(years=int(minimum_maximum_age[0]))
                    min_dob = date_of_enrollment - relativedelta(years=int(minimum_maximum_age[1]))

                    return render(request, 'client_baseline_data.html', {'page': 'clients',
                                                                         'page_title': 'DREAMS Enrollment Data',
                                                                         'client': client_found,
                                                                         'user': request.user,
                                                                         'demo_form': demographics_form,
                                                                         'household_form': household_form,
                                                                         'edu_form': edu_and_emp_form,
                                                                         'sexuality_form': sexuality_form,
                                                                         'gbv_form': gbv_form,
                                                                         'hiv_form': hiv_form,
                                                                         'rh_form': reproductive_health_form,
                                                                         'drug_use_form': drug_use_form,
                                                                         'programe_participation_form': participation_form,
                                                                         'search_client_term': search_client_term,
                                                                         'transfer_form': ClientTransferForm(
                                                                             ip_code=ip_code,
                                                                             initial={'client': client_found}),
                                                                         'is_editable_by_ip': is_editable_by_ip,
                                                                         'client_status': client_status,
                                                                         'max_dob': max_dob,
                                                                         'min_dob': min_dob
                                                                         })
            except Client.DoesNotExist:
                traceback.format_exc()
                return redirect('clients')
            except Exception as e:
                traceback.format_exc()
                return redirect('clients')
    else:
        return redirect('login')


def update_demographics_data(request):
    client_id = int(request.POST['client'], 0)
    instance = Client.objects.get(id=client_id)
    if request.is_ajax():
        if request.method == 'POST':
            county_of_residence = instance.county_of_residence
            sub_county = instance.sub_county
            implementing_partner = instance.implementing_partner
            dreams_id = instance.dreams_id
            form = DemographicsForm(request.POST, instance=instance)

            if form.is_valid():
                client_enrolment_service_layer = ClientEnrolmentServiceLayer(request.user)
                if not client_enrolment_service_layer.is_within_enrolment_dates(form.instance.date_of_birth,
                                                                                form.instance.date_of_enrollment):
                    min_max_age = client_enrolment_service_layer.get_minimum_maximum_enrolment_age(
                        client_enrolment_service_layer.ENROLMENT_CUTOFF_DATE)

                    response_data = {
                        'status': 'fail',
                        'errors': [
                            "The client is not within the accepted age range. At the date of enrolment the age of the client must be between " + str(
                                min_max_age[0]) + " and " + str(min_max_age[1] + " years.")],
                        'client_age': instance.get_current_age()
                    }
                    return JsonResponse(response_data, status=500)

                form.instance.implementing_partner = implementing_partner
                form.instance.county_of_residence = county_of_residence
                form.instance.sub_county = sub_county
                form.instance.implementing_partner = implementing_partner
                form.instance.dreams_id = dreams_id
                form.instance.save(user_id=request.user.id, action="UPDATE")

                response_data = {
                    'status': 'success',
                    'errors': form.errors,
                    'client_age': instance.get_current_age()
                }
                return JsonResponse(response_data, status=200)
            else:
                response_data = {
                    'status': 'fail',
                    'errors': form.errors
                }
                return JsonResponse(response_data, status=500)
        else:
            raise PermissionDenied
    else:
        raise PermissionDenied
    return render(request, template, {'status': 'success'})


def update_individual_and_household_data(request):
    client_id = int(request.POST['client'])
    instance = ClientIndividualAndHouseholdData.objects.get(client=client_id)
    if request.is_ajax():
        template = 'ajax_response_form/client_individual_household_ajax_form.html'

        if request.method == 'POST':
            form = IndividualAndHouseholdForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                response_data = {
                    'status': 'success'
                }
                return JsonResponse(response_data, status=200)
            else:
                response_data = {
                    'status': 'fail',
                    'errors': form.errors
                }
                return JsonResponse(response_data, status=500)
        else:
            raise PermissionDenied
    else:
        raise PermissionDenied
    return render(request, template, {'household_form': form})


def update_edu_and_employment_data(request):
    client_id = int(request.POST['client'])
    instance = ClientEducationAndEmploymentData.objects.get(client=client_id)
    if request.is_ajax():
        template = 'ajax_response_form/education_and_employment_ajax_form.html'
        form = EducationAndEmploymentForm(request.POST, instance=instance)
        if request.method == 'POST':
            if form.is_valid():
                form.save()
                response_data = {
                    'status': 'success'
                }
                return JsonResponse(response_data, status=200)
            else:
                print (form.errors)
        else:
            response_data = {
                'status': 'fail',
                'errors': form.errors
            }
            return JsonResponse(response_data, status=500)
    else:
        raise PermissionDenied
    return render(request, template, {'edu_form': form})


def update_hiv_testing_data(request):
    client_id = int(request.POST['client'])
    instance = ClientHIVTestingData.objects.get(client=client_id)
    if request.is_ajax():
        template = 'ajax_response_form/client_hiv_testing_ajax_form.html'

        if request.method == 'POST':
            form = HivTestForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                response_data = {
                    'status': 'success'
                }
                return JsonResponse(response_data, status=200)
            else:
                response_data = {
                    'status': 'fail',
                    'errors': form.errors
                }
                return JsonResponse(response_data, status=500)
        else:
            raise PermissionDenied
    else:
        raise PermissionDenied
    return render(request, template, {'hiv_form': form})


def update_sexuality_data(request):
    client_id = int(request.POST['client'])
    instance = ClientSexualActivityData.objects.get(client=client_id)
    if request.is_ajax():
        if request.method == 'POST':
            form = SexualityForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                response_data = {
                    'status': 'success'
                }
                return JsonResponse(response_data, status=200)
            else:
                response_data = {
                    'status': 'fail',
                    'errors': form.errors
                }
                return JsonResponse(response_data, status=500)
        else:
            raise PermissionDenied
    else:
        raise PermissionDenied
    return render(request, template, {'sexuality_form': form})


def update_rep_health_data(request):
    client_id = int(request.POST['client'])
    instance = ClientReproductiveHealthData.objects.get(client=client_id)
    if request.is_ajax():
        template = 'ajax_response_form/client_reproductive_health_ajax_form.html'

        if request.method == 'POST':
            form = ReproductiveHealthForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                response_data = {
                    'status': 'success'
                }
                return JsonResponse(response_data, status=200)
            else:
                response_data = {
                    'status': 'fail',
                    'errors': form.errors
                }
                return JsonResponse(response_data, status=500)
        else:
            raise PermissionDenied
    else:
        raise PermissionDenied
    return render(request, template, {'rh_form': form})


def update_gbv_data(request):
    client_id = int(request.POST['client'])
    instance = ClientGenderBasedViolenceData.objects.get(client=client_id)
    if request.is_ajax():
        template = 'ajax_response_form/client_gbv_ajax_form.html'

        if request.method == 'POST':
            form = GBVForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                response_data = {
                    'status': 'success'
                }
                return JsonResponse(response_data, status=200)
            else:
                response_data = {
                    'status': 'fail',
                    'errors': form.errors
                }
                return JsonResponse(response_data, status=500)
        else:
            raise PermissionDenied
    else:
        raise PermissionDenied
    return render(request, template, {'gbv_form': form})


def update_drug_use_data(request):
    client_id = int(request.POST['client'])
    instance = ClientDrugUseData.objects.get(client=client_id)
    if request.is_ajax():
        template = 'ajax_response_form/client_drug_use_ajax_form.html'

        if request.method == 'POST':
            form = DrugUseForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                response_data = {
                    'status': 'success'
                }
                return JsonResponse(response_data, status=200)
            else:
                response_data = {
                    'status': 'fail',
                    'errors': form.errors
                }
                return JsonResponse(response_data, status=500)
        else:
            raise PermissionDenied
    else:
        raise PermissionDenied
    return render(request, template, {'drug_use_form': form})


def update_programme_participation_data(request):
    client_id = int(request.POST['client'])
    instance = ClientParticipationInDreams.objects.get(client=client_id)
    if request.is_ajax():
        template = 'ajax_response_form/client_programme_participation_ajax_form.html'

        if request.method == 'POST':
            form = DreamsProgramParticipationForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                response_data = {
                    'status': 'success'
                }
                return JsonResponse(response_data, status=200)
            else:
                response_data = {
                    'status': 'fail',
                    'errors': form.errors
                }
                return JsonResponse(response_data, status=500)
        else:
            raise PermissionDenied
    else:
        raise PermissionDenied
    return render(request, template, {'programe_participation_form': form})


def transfer_client(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if request.method == 'POST' and request.is_ajax():
                try:
                    ip = request.user.implementingpartneruser.implementing_partner
                except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
                    response_data = {
                        'status': 'fail',
                        'message': 'Transfer Failed. You do not belong to an implementing partner',
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)

                transfer_form = ClientTransferForm(request.POST)
                if transfer_form.is_valid():
                    num_of_pending_transfers = ClientTransfer.objects.filter(client=transfer_form.instance.client,
                                                                             transfer_status=ClientTransferStatus.objects.get(
                                                                                 code__exact=TRANSFER_INITIATED_STATUS)).count()

                    if num_of_pending_transfers > 0:
                        print ("{} pending transfers for client".format(num_of_pending_transfers))
                        response_data = {
                            'status': 'fail',
                            'message': "Transfer failed, there's a pending transfer for this client.",
                        }
                    else:
                        client_transfer = transfer_form.save(commit=False)

                        client_transfer.transfer_status = ClientTransferStatus.objects.get(code__exact=TRANSFER_INITIATED_STATUS)
                        client_transfer.source_implementing_partner = ip
                        client_transfer.initiated_by = request.user
                        client_transfer.start_date = dt.now()
                        client_transfer.save()

                        client = transfer_form.instance.client
                        client.date_changed = dt.now()
                        client.save()

                        response_data = {
                            'status': 'success',
                            'message': 'Transfer request received, pending approval by the receiving implementing partner.',
                        }
                    return JsonResponse(json.dumps(response_data), safe=False)
                else:
                    response_data = {
                        'status': 'fail',
                        'message': transfer_form.errors,
                    }
                    return JsonResponse(json.dumps(response_data), safe=False)
        else:
            raise PermissionDenied
    except Exception as e:
        response_data = {
            'status': 'fail',
            'message': str(e),
        }
        return JsonResponse(json.dumps(response_data), safe=False)


def client_transfers(request, *args, **kwargs):
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:
        transferred_in = bool(int(kwargs.pop('transferred_in', 1)))

        transfer_perm = TransferServiceLayer(request.user)
        can_accept_or_reject = transfer_perm.can_accept_or_reject_transfer()

        try:
            ip = request.user.implementingpartneruser.implementing_partner
            if transferred_in:
                c_transfers = ClientTransfer.objects.filter(destination_implementing_partner=ip).order_by('transfer_status', '-date_created')
            else:
                c_transfers = ClientTransfer.objects.filter(source_implementing_partner=ip).order_by('transfer_status', '-date_created')

        except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
            c_transfers = ClientTransfer.objects.all()

        page = request.GET.get('page', 1)
        paginator = Paginator(c_transfers, 20)

        try:
            transfers = paginator.page(page)
        except PageNotAnInteger:
            transfers = paginator.page(1)
        except EmptyPage:
            transfers = paginator.page(paginator.num_pages)

        return render(request, "client_transfers.html",
                      {'client_transfers': transfers, 'can_accept_or_reject': can_accept_or_reject,
                       'transferred_in': transferred_in, 'page': 'transfers'})
    else:
        return redirect('login')


def client_referrals(request, *args, **kwargs):
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:
        referred_in = bool(int(kwargs.pop('referred_in', 1)))

        refer_perm = ReferralServiceLayer(request.user)
        can_accept_or_reject = refer_perm.can_accept_or_reject_referral()

        try:
            ip = request.user.implementingpartneruser.implementing_partner
            if referred_in:
                c_referrals = Referral.objects.filter(receiving_ip=ip).order_by('referral_status', '-referral_date')
            else:
                c_referrals = Referral.objects.filter(referring_ip=ip).order_by('referral_status', '-referral_date')

        except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
            return render(request, 'login.html')

        page = request.GET.get('page', 1)
        paginator = Paginator(c_referrals, 20)

        try:
            referrals = paginator.page(page)
        except PageNotAnInteger:
            referrals = paginator.page(1)
        except EmptyPage:
            referrals = paginator.page(paginator.num_pages)

        return render(request, "client_referrals.html",
                      {'client_referrals': referrals, 'can_accept_or_reject': can_accept_or_reject,
                       'referred_in': referred_in, 'page': 'referrals'})
    else:
        return redirect('login')


def accept_client_transfer(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if request.method == 'POST':

                try:
                    ip = request.user.implementingpartneruser.implementing_partner
                except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
                    if not request.user.is_superuser:
                        messages.error(request, "You do not belong to an implementing partner")
                        return redirect(reverse("client_transfers", kwargs={'transferred_in': 1}))
                    else:
                        ip = None

                client_transfer_id = request.POST.get("id", "")
                if client_transfer_id != "":
                    client_transfer = ClientTransfer.objects.get(id__exact=client_transfer_id)

                    transfer_perm = TransferServiceLayer(request.user, client_transfer=client_transfer)
                    can_accept_transfer = transfer_perm.can_accept_transfer()

                    if not can_accept_transfer:
                        raise PermissionDenied

                    if client_transfer is not None:
                        accepted_client_transfer_status = ClientTransferStatus.objects.get(code__exact=TRANSFER_ACCEPTED_STATUS)

                        client_transfer.transfer_status = accepted_client_transfer_status
                        client_transfer.completed_by = request.user
                        client_transfer.end_date = dt.now()

                        # Update the client to receive interventions from this new ip.
                        client = Client.objects.get(id__exact=client_transfer.client.id)
                        if ip is None and request.user.is_superuser:
                            ip = client_transfer.destination_implementing_partner
                        client.implementing_partner = ip

                        with transaction.atomic():
                            client.save()
                            client_transfer.save()

                        messages.info(request, "Transfer successfully accepted.")
                        return redirect("/client?client_id={}".format(client.id))
                    else:
                        messages.error(request,
                                       "Transfer not effected. Contact System Administrator if this error Persists.")
                else:
                    messages.error(request,
                                   "Transfer not effected. Contact System Administrator if this error Persists.")
        else:
            raise PermissionDenied
    except Exception as e:
        messages.error(request,
                       "An error occurred while processing request. "
                       "Contact System Administrator if this error Persists.")

    return redirect(reverse("client_transfers", kwargs={'transferred_in': 1}))


def reject_client_transfer(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if request.method == 'POST':

                client_transfer_id = request.POST.get("id", "")
                if client_transfer_id != "":
                    client_transfer = ClientTransfer.objects.get(id__exact=client_transfer_id)
                else:
                    client_transfer = None

                if client_transfer is not None:
                    transfer_perm = TransferServiceLayer(request.user, client_transfer=client_transfer)
                    can_reject_transfer = transfer_perm.can_reject_transfer()

                    if not can_reject_transfer:
                        raise PermissionDenied

                    client_transfer.transfer_status = ClientTransferStatus.objects.get(code__exact=TRANSFER_REJECTED_STATUS)
                    client_transfer.completed_by = request.user
                    client_transfer.end_date = dt.now()
                    client_transfer.save()

                    messages.info(request, "Transfer successfully rejected.")
                else:
                    messages.warning(request,
                                     "Transfer not rejected. Contact System Administrator if this error Persists.")
        else:
            raise PermissionDenied
    except Exception as e:
        messages.error(request,
                       "An error occurred while processing request. "
                       "Contact System Administrator if this error Persists.")

    return redirect(reverse("client_transfers", kwargs={'transferred_in': 1}))


def get_client_transfers_count(request):
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:
        initiated_client_transfer_status = ClientTransferStatus.objects.get(code__exact=TRANSFER_INITIATED_STATUS)
        try:
            ip = request.user.implementingpartneruser.implementing_partner
            client_transfers_count = ClientTransfer.objects.filter(
                destination_implementing_partner=ip,
                transfer_status=initiated_client_transfer_status).count()
        except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
            client_transfers_count = ClientTransfer.objects.filter(
                transfer_status=initiated_client_transfer_status).count()
        except Exception:
            client_transfers_count = 0

        return HttpResponse(client_transfers_count)
    else:
        return HttpResponse(0)


def get_client_referrals_count(request):
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:
        initiated_referral_status = ReferralStatus.objects.get(code__exact=ReferralServiceLayer.REFERRAL_PENDING_STATUS)
        try:
            ip = request.user.implementingpartneruser.implementing_partner
            client_referrals_count = Referral.objects.filter(
                receiving_ip=ip,
                referral_status=initiated_referral_status).count()
        except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
            client_referrals_count = 0
        except Exception:
            client_referrals_count = 0

        return HttpResponse(client_referrals_count)
    else:
        return HttpResponse(0)


def intervention_export_transferred_in_page(request):
    if request.user.is_authenticated() and request.user.is_active and request.user.has_perm(
            'DreamsApp.can_export_raw_data'):

        try:
            ips = None
            if request.user.is_superuser or request.user.has_perm('DreamsApp.can_view_cross_ip_data'):
                ips = ImplementingPartner.objects.all()
            elif request.user.implementingpartneruser is not None:
                ips = ImplementingPartner.objects.filter(
                    id=request.user.implementingpartneruser.implementing_partner.id)
            if ips.count() > 0:
                ips = ips.union(ImplementingPartner.objects.filter(parent_implementing_partner__in=ips))

            context = {'page': 'export', 'page_title': 'Interventions Transferred In Export', 'ips': ips,
                       'counties': County.objects.all()}
            return render(request, 'interventionDataExportTransferredIn.html', context)

        except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
            traceback.format_exc()
    else:
        raise PermissionDenied


def download_raw_intervention_transferred_in_export(request):
    try:
        from_intervention_date = request.POST.get('from_intervention_date')
        to_intervention_date = request.POST.get('to_intervention_date')

        export_file_name = urllib.parse.quote(
            ("/tmp/raw_intervention_transferred_in_export-{}.csv").format(datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        export_doc = DreamsRawExportTemplateRenderer()

        if request.user.is_superuser or request.user.has_perm('DreamsApp.can_view_phi_data') \
                or Permission.objects.filter(group__user=request.user).filter(
            codename='DreamsApp.can_view_phi_data').exists():
            show_PHI = True
        else:
            show_PHI = False

        try:
            ip = request.user.implementingpartneruser.implementing_partner
        except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
            ip = None

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = ('attachment; filename="{}"').format(export_file_name)
        export_doc.get_intervention_excel_transferred_in_doc(response, ip, from_intervention_date, to_intervention_date,
                                                             show_PHI)
        return response

    except Exception as e:
        raise e


def export_client_transfers(request, *args, **kwargs):
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:

        transferred_in = bool(int(kwargs.pop('transferred_in', 1)))
        columns = ("client__dreams_id", "source_implementing_partner__name",
                   "destination_implementing_partner__name", "transfer_reason", "transfer_status__name",)

        try:
            ip = request.user.implementingpartneruser.implementing_partner
            if transferred_in:
                c_transfers = ClientTransfer.objects.values_list(*columns).filter(destination_implementing_partner=ip)
            else:
                c_transfers = ClientTransfer.objects.values_list(*columns).filter(source_implementing_partner=ip)
        except (ImplementingPartnerUser.DoesNotExist, ImplementingPartner.DoesNotExist):
            c_transfers = ClientTransfer.objects.values_list(*columns)

        header = ['Dreams ID', 'Source Implementing Partner', 'Destination Implementing Partner', 'Transfer Reason',
                  'Status']

        wb = Workbook()
        ws = wb.active
        ws.append(header)

        for c_transfer in c_transfers:
            ws.append(c_transfer)

        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))

        for col, value in dims.items():
            ws.column_dimensions[col].width = value

        ft = Font(bold=True)
        for cell in ws["1:1"]:
            cell.font = ft

        file_name = "Client_Transfers_{}.xlsx".format("In" if transferred_in else "Out")
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename={}'.format(file_name)

        wb.save(response)
        return response
    else:
        return redirect('login')


def export_client_referrals(request, *args, **kwargs):
    if request.user is not None and request.user.is_authenticated() and request.user.is_active:

        return None
    else:
        return redirect('login')


def void_client(request):
    try:
        if request.user is not None and request.user.is_authenticated() and request.user.is_active:
            if request.method == 'POST' and request.is_ajax():
                if not request.user.is_superuser and not request.user.has_perm(
                        'DreamsApp.can_void_client') and not Permission.objects.filter(group__user=request.user).filter(
                        codename='DreamsApp.can_void_client').exists():
                    return get_response_data(0, 'Voiding Failed. You do not have permission to void a client')

                client_id = request.POST.get("id", "")
                void_reason = request.POST.get("void_reason", "")
                if void_reason is None or void_reason == "" or void_reason.isspace():
                    return get_response_data(0, {'void_reason': ['Reason for voiding is required.']})

                if client_id != "":
                    cursor = db_conn_2.cursor()
                    try:
                        args = [client_id, request.user.id, void_reason]
                        cursor.callproc('sp_void_client_by_id', args)
                        exec_status = cursor.fetchone()[0]

                        if exec_status == 1:
                            messages.info(request, "Client has been voided.")
                            response_data = get_response_data(1, 'Client has been voided.', next_url=reverse('clients'))
                        else:
                            response_data = get_response_data(0,
                                                              'Client not voided, please contact system '
                                                              'administrator for assistance.')
                    except Exception as e:
                        response_data = get_response_data(0, 'Client could not be voided.{}'.format(e))
                    finally:
                        cursor.close()

                    return response_data
                else:
                    return get_response_data(0, 'Invalid client ID. Please specify client.')
            else:
                raise SuspiciousOperation
        else:
            raise PermissionDenied
    except Exception as e:
        traceback.format_exc()
        return get_response_data(0, e)


def get_response_data(status, message, **kwargs):
    response = {
        'status': 'success' if status == 1 else 'fail',
        'message': message
    }

    for k, v in kwargs.items():
        response[k] = v

    return JsonResponse(json.dumps(response), safe=False)


def download_audit_logs(request):
    if not request.user.is_superuser and not request.user.has_perm('DreamsApp.can_manage_audit'):
        raise PermissionDenied('Operation not allowed. [Missing Permission]')

    ip = ''
    try:
        ip = request.user.implementingpartneruser.implementing_partner.id
    except ImplementingPartnerUser.DoesNotExist:
        pass

    # user is allowed to view logs
    if request.method == 'GET':
        try:
            filter_text = request.GET.get('filter-log-text', '')
            filter_date_from = request.GET.get('filter-log-date-from', '')
            filter_date = request.GET.get('filter-log-date', '')

            # getting logs
            logs = Audit.objects.filter(Q(table__in=filter_text.split(" ")) |
                                        Q(action__in=filter_text.split(" ")) |
                                        Q(search_text__in=filter_text.split(" ")) |
                                        Q(user__username__icontains=filter_text.split(" ")[0]) |
                                        Q(user__first_name__icontains=filter_text.split(" ")[0]) |
                                        Q(user__last_name__icontains=filter_text.split(" ")[0])
                                        ).order_by('-timestamp')

            logs = filter_audit_logs_by_date_and_ip(filter_date, filter_date_from, ip, logs)

            header = ['Timestamp', 'User', 'Table', 'Field', 'Old Value', 'New Value', 'Action', 'Text']

            wb = Workbook()
            ws = wb.active
            ws.append(header)
            row_idx = 2

            for log in logs:
                ws.cell(row=row_idx, column=1).value = log.timestamp
                ws.cell(row=row_idx, column=2).value = log.get_user_name()
                ws.cell(row=row_idx, column=3).value = log.table

                adt_idx = row_idx
                for audittrail in log.audittrail_set.all():
                    ws.cell(row=adt_idx, column=4).value = audittrail.column
                    ws.cell(row=adt_idx, column=5).value = audittrail.old_value
                    ws.cell(row=adt_idx, column=6).value = audittrail.new_value
                    adt_idx += 1

                if adt_idx == row_idx:
                    adt_idx += 1

                ws.cell(row=row_idx, column=7).value = log.action
                ws.cell(row=row_idx, column=8).value = log.search_text
                row_idx = adt_idx

            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column] = max(dims.get(cell.column, 0), len(str(cell.value)))

            for col, value in dims.items():
                ws.column_dimensions[col].width = value

            ft = Font(bold=True)
            for cell in ws["1:1"]:
                cell.font = ft

            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename=Audit_Logs.xlsx'

            wb.save(response)
            return response

        except Exception as e:
            tb = traceback.format_exc(e)
            return HttpResponseServerError(tb)
    else:
        raise SuspiciousOperation


def get_min_max_date_of_birth(request):
    try:
        if request.method == 'POST' and request.user is not None and request.user.is_authenticated() and request.user.is_active:
            date_of_enrollment_str = request.POST.get('date_of_enrollment')
            date_of_enrollment = datetime.strptime(str(date_of_enrollment_str),
                                                   '%Y-%m-%d').date() if date_of_enrollment_str is not None else dt.now().date()
            client_enrolment_service_layer = ClientEnrolmentServiceLayer(request.user)
            minimum_maximum_age = client_enrolment_service_layer.get_minimum_maximum_enrolment_age(
                client_enrolment_service_layer.ENROLMENT_CUTOFF_DATE)
            max_dob = date_of_enrollment - relativedelta(years=int(minimum_maximum_age[0]))
            min_dob = date_of_enrollment - relativedelta(years=int(minimum_maximum_age[1]))

            response_data = {
                "min_dob": min_dob,
                "max_dob": max_dob
            }
            return JsonResponse(response_data)

        else:
            raise PermissionDenied
    except Exception as e:
        tb = traceback.format_exc(e)
        return HttpResponseServerError(tb)
