from django.core.files.storage import default_storage
import openpyxl as xl
import xlrd
import traceback
from Dreams_Excel_Mapping import *
from string import Template
import datetime
from django.db import connection
from django.conf import settings
from openpyxl.utils.exceptions import *
import os
import csv


class DreamsEnrollmentExcelTemplateRenderer(object):

    document_path = ''
    workbook = ''

    def __init__(self):
        pass

    def create_tmp_file(self, excel):
        try:
            with open(default_storage.path('tmp/enrollment.xlsx'), 'wb+') as destination:
                for chunk in excel.chunks():
                    destination.write(chunk)
            destination.closed
            self.document_path = default_storage.path('tmp/enrollment.xlsx')
            return default_storage.path('tmp/enrollment.xlsx')
        except:
            print 'Could not create a temp file'
            return

    def get_document_path(self):
        return self.document_path

    def validate_excel_document(self):
        # check to see if Main Database sheet exists
        # TODO define and implement more checks

        wb = xl.load_workbook(self.get_document_path(), read_only=True, keep_vba=False)
        self.workbook = wb
        if 'Main Database' in wb.get_sheet_names():
            return True
        else:
            return False

    def get_sheet_names(self):
        if self.validate_excel_document():
            return xl.load_workbook(self.get_document_path(), read_only=True, keep_vba=False).get_sheet_names()
        else:
            return 'Invalid Sheets'

    def excel_enrollment_data(self):
        if self.validate_excel_document():
            try:
                sql_statement = self.generate_insert_SQL(6, 50)
                # print sql_statement
                self.execute_SQL_Query(sql_statement)

            except Exception as e:
                tb = traceback.format_exc()
                print 'There was an error ' + tb
            return

    def generate_insert_SQL(self, start, end):
        if self.validate_excel_document():
            try:
                print 'The Excel Database is Valid'
                print "Now processing the file.........."

                sql_statement = """
INSERT INTO DreamsApp_client
(first_name, middle_name, last_name, dreams_id, dss_id_number, implementing_partner_id, date_of_birth, date_of_enrollment, marital_status_id, phone_number,
county_of_residence_id, sub_county_id, ward_id, informal_settlement, village, landmark, guardian_name, relationship_with_guardian, guardian_phone_number, guardian_national_id, is_date_of_birth_estimated, verification_document_id, verification_doc_no )
VALUES """

                rangeList = range(start, end)
                lastNumber = rangeList[-1]
                row_qry = ''
                for rowid in range(start, end):

                    if rowid != lastNumber:
                        row_qry += self.generate_row_insert_SQL(rowid) + ', \n'
                    else:
                        row_qry += " " + self.generate_row_insert_SQL(rowid)

                return Template('$insert $values').substitute(insert=sql_statement, values=row_qry)

            except Exception as e:
                tb = traceback.format_exc()
                print 'There was an error ' + tb
            return

    def get_enrollment_sheet_openpyxl(self):
        return xl.load_workbook(self.get_document_path(), read_only=True, data_only=True).get_sheet_by_name('Main Database')

    def get_enrollment_sheet(self):
        return self.get_wb().sheet_by_name('Main Database')

    def get_wb(self):
        return xlrd.open_workbook(self.get_document_path(), on_demand=True)

    def get_row_data(self, row_num):

        sheet = self.get_enrollment_sheet_openpyxl()
        cols = self.openpyxl_demographic_columns()

        # map fields that should mirror webapp's

        ver_doc = str(sheet.cell(row=row_num, column=cols.get('verification_doc')).value)
        ver_doc = ExcelDreamsMapping.verification_document().get(ver_doc)

        marital_status = str(sheet.cell(row=row_num, column=cols.get('marital_status')).value)
        marital_status = ExcelDreamsMapping.marital_status_codes().get(marital_status)

        county = str(sheet.cell(row=row_num, column=cols.get('county')).value)
        county = ExcelDreamsMapping.county().get(county)

        sub_county = str(sheet.cell(row=row_num, column=cols.get('sub_county')).value)
        sub_county = ExcelDreamsMapping.sub_county().get(sub_county)

        ward = str(sheet.cell(row=row_num, column=cols.get('ward')).value)
        ward = ExcelDreamsMapping.ward_by_name().get(ward)

        dob_raw = sheet.cell(row=row_num, column=cols.get('dob')).value
        if dob_raw is not None:
            dob_raw = datetime.datetime.strftime(dob_raw, '%Y-%m-%d')

        doe_raw = sheet.cell(row=row_num, column=cols.get('date_of_enrollment')).value
        if doe_raw is not None:
            doe_raw = datetime.datetime.strftime(doe_raw, '%Y-%m-%d')

        row_values = {
            'serial_No': str(sheet.cell(row=row_num, column=cols.get('serial_No')).value),
            'IP': str(sheet.cell(row=row_num, column=cols.get('IP')).value),
            'first_name': str(sheet.cell(row=row_num, column=cols.get('first_name')).value),
            'middle_name': str(sheet.cell(row=row_num, column=cols.get('middle_name')).value),
            'last_name': str(sheet.cell(row=row_num, column=cols.get('last_name')).value),
            'dob': dob_raw,
            'verification_doc': ver_doc,
            'verification_doc_other': str(sheet.cell(row=row_num, column=cols.get('verification_doc_other')).value),
            'verification_doc_no': str(sheet.cell(row=row_num, column=cols.get('verification_doc_no')).value),
            'date_of_enrollment': doe_raw,
            'marital_status': marital_status,
            'client_phone_no': str(sheet.cell(row=row_num, column=cols.get('client_phone_no')).value),
            'county': county,
            'sub_county': sub_county,
            'ward': ward,
            'informal_settlement': str(sheet.cell(row=row_num, column=cols.get('informal_settlement')).value),
            'village': str(sheet.cell(row=row_num, column=cols.get('village')).value),
            'land_mark': str(sheet.cell(row=row_num, column=cols.get('land_mark')).value),
            'dreams_id': str(sheet.cell(row=row_num, column=cols.get('dreams_id')).value),
            'dss_id': str(sheet.cell(row=row_num, column=cols.get('dss_id')).value),
            'caregiver_first_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_first_name')).value),
            'caregiver_middle_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_middle_name')).value),
            'caregiver_last_name': str(sheet.cell(row=row_num, column=cols.get('caregiver_last_name')).value),
            'caregiver_relationship': str(sheet.cell(row=row_num, column=cols.get('caregiver_relationship')).value),
            'caregiver_relationship_other': str(sheet.cell(row=row_num, column=cols.get('caregiver_relationship_other')).value),
            'caregiver_phone_no': str(sheet.cell(row=row_num, column=cols.get('caregiver_phone_no')).value),
            'caregiver_id_no': str(sheet.cell(row=row_num, column=cols.get('caregiver_id_no')).value),
        }
        return row_values


    def generate_row_insert_SQL(self, row_num):

        row_data = self.get_row_data(row_num)
        caregiver_fname = row_data.get('caregiver_first_name')
        caregiver_mname = row_data.get('caregiver_middle_name')
        caregiver_lname = row_data.get('caregiver_last_name')

        coded_rel = row_data.get('caregiver_relationship')
        other_rel = row_data.get('caregiver_relationship_other')
        date_of_birth_estimated = 0

        # django app has no field for other relationship. check which one has value

        if coded_rel != 'None' or other_rel != 'None':
            if other_rel != 'None':
                caregiver_relationship = other_rel
            else:
                caregiver_relationship = coded_rel
        else:
            caregiver_relationship = ''

        ver_doc = row_data.get('verification_doc')
        ver_doc_other = row_data.get('verification_doc_other')

        # TODO: Handle other verification documents. It is not currently in the Client's model. Code therefore picks only coded documents

        guardian_name = (caregiver_fname + " " if caregiver_fname != 'None' else "") + (caregiver_mname + " " if caregiver_mname != 'None' else "") + (caregiver_lname if caregiver_lname != 'None' else "")

        values_template = Template(
            '("$first_name", "$middle_name", "$last_name", "$dreams_id", "$dss_id_number", "$implementing_partner_id", "$date_of_birth", "$date_of_enrollment", "$marital_status_id", "$phone_number", \
    "$county_of_residence_id", "$sub_county_id", "$ward_id", "$informal_settlement", "$village", "$landmark", "$guardian_name", "$relationship_with_guardian", "$guardian_phone_number", "$guardian_national_id", "$is_date_of_birth_estimated", "$verification_doc_id", "$verification_doc_no" )').safe_substitute(
            first_name=row_data.get('first_name') if row_data.get('first_name') !='None' else '',
            middle_name=row_data.get('middle_name') if row_data.get('middle_name') !='None' else '',
            last_name=row_data.get('last_name') if row_data.get('last_name') !='None' else '',
            dreams_id=row_data.get('dreams_id'),
            dss_id_number=row_data.get('dss_id') if row_data.get('dss_id') !='None' else '',
            implementing_partner_id=row_data.get('IP') if row_data.get('IP') !='None' else '',
            date_of_birth=row_data.get('dob') if row_data.get('dob') !='None' else '0000-00-00',
            date_of_enrollment=row_data.get('date_of_enrollment') if row_data.get('date_of_enrollment') !='None' else '0000-00-00',
            marital_status_id=row_data.get('marital_status') if row_data.get('marital_status') !='None' else '',
            phone_number=row_data.get('client_phone_no') if row_data.get('client_phone_no') !='None' else '',
            county_of_residence_id=row_data.get('county'),
            sub_county_id=row_data.get('sub_county'),
            ward_id=row_data.get('ward'),
            informal_settlement=row_data.get('informal_settlement') if row_data.get('informal_settlement') !='None' else '',
            village=row_data.get('village') if row_data.get('village') !='None' else '',
            landmark=row_data.get('land_mark') if row_data.get('land_mark') !='None' else '',
            guardian_name=guardian_name,
            relationship_with_guardian=caregiver_relationship,
            guardian_phone_number=row_data.get('caregiver_phone_no') if row_data.get('caregiver_phone_no') !='None' else '',
            guardian_national_id=row_data.get('caregiver_id_no') if row_data.get('caregiver_id_no') !='None' else '',
            is_date_of_birth_estimated=date_of_birth_estimated,
            verification_doc_id=row_data.get('verification_doc'),
            verification_doc_no=row_data.get('verification_doc_no')
        )

        return values_template

    def openpyxl_demographic_columns(self):
        demographics = {
            'serial_No': 1,
            'IP': 3,
            'first_name': 4,
            'middle_name': 5,
            'last_name': 6,
            'dob': 7,
            'verification_doc': 8,
            'verification_doc_other': 9,
            'verification_doc_no': 10,
            'date_of_enrollment': 11,
            'marital_status': 16,
            'client_phone_no': 17,
            'county': 18,
            'sub_county': 19,
            'ward': 20,
            'informal_settlement': 22,
            'village': 23,
            'land_mark': 24,
            'dreams_id': 25,
            'dss_id': 26,
            'caregiver_first_name': 27,
            'caregiver_middle_name': 28,
            'caregiver_last_name': 29,
            'caregiver_relationship': 30,
            'caregiver_relationship_other': 31,
            'caregiver_phone_no': 32,
            'caregiver_id_no': 33,
        }

        return demographics

    def execute_SQL_Query(self, sql):
        cursor = connection.cursor()
        try:
            cursor.execute(sql)

            print cursor.rowcount
        except Exception as e:
            print 'There was an Error running the query\n'
            traceback.format_exc()

        return

    def get_export_rows(self, ip_list_str, sub_county, ward):
        cursor = connection.cursor()
        multiple_ip_sub_county_query = "SELECT * FROM flat_dreams_enrollment WHERE voided=0 AND sub_county_code = %s AND  implementing_partner_id IN %s "
        multiple_ip_ward_query = "SELECT * FROM flat_dreams_enrollment WHERE voided=0 AND ward_id = %s AND  implementing_partner_id IN %s "
        multiple_ip_default_query = "SELECT * FROM flat_dreams_enrollment WHERE voided=0 AND implementing_partner_id IN %s "

        single_ip_sub_county_query = "SELECT * FROM flat_dreams_enrollment WHERE voided=0 AND sub_county_code = %s AND implementing_partner_id = %s "
        single_ip_ward_query = "SELECT * FROM flat_dreams_enrollment WHERE voided=0 AND ward_id = %s AND implementing_partner_id = %s "
        single_ip_default_query = "SELECT * FROM flat_dreams_enrollment WHERE voided=0 AND implementing_partner_id = %s "

        cursor_results = None

        try:

            ip_tuple_l = ip_list_str
            if sub_county is not None and sub_county:
                sub_county = int(sub_county)

            if ward is not None and ward:
                ward = int(ward)

            if len(ip_tuple_l) > 1:
                ip_list = tuple(ip_tuple_l)

                if ward is not None and ward:
                    cursor.execute(multiple_ip_ward_query, [ward, ip_list])
                elif sub_county is not None and sub_county:
                    cursor.execute(
                        multiple_ip_sub_county_query, [sub_county, ip_list])
                else:
                    cursor.execute(multiple_ip_default_query, [ip_list])
            else:
                ip_list = ip_list_str[0]
                if ward is not None and ward:
                    cursor.execute(single_ip_ward_query, [ward, ip_list])
                elif sub_county is not None and sub_county:
                    cursor.execute(single_ip_sub_county_query, [sub_county, ip_list])
                else:
                    cursor_results = cursor.execute(single_ip_default_query, [ip_list])

            print "Query was successful"

            return cursor_results, cursor

        except Exception as e:
            print 'There was an Error running the query\n'
            traceback.format_exc()

        return

    def fetch_intervention_rows(self, ip_list_str, sub_county, ward):
        cursor = connection.cursor()

        multiple_ip_sub_county_query = """select
  i.client_id, i.dreams_id, CONCAT_WS(" ",i.first_name, i.middle_name, i.last_name) AS client_name, i.date_of_birth, i.implementing_partner, i.implementing_partner_id, i.county_of_residence,i.sub_county,
  i.ward, i.village, i.date_of_enrollment as date_of_enrollment, DATE(i.intervention_date) date_of_intervention, DATE(i.date_created) date_created, i.intervention as intervention_type, i.intervention_category, i.hts_result,
  i.pregnancy_test_result, i.client_ccc_number, i.date_linked_to_ccc,
  i.no_of_sessions_attended, i.comment, i.current_age, i.age_at_intervention
from stag_client_intervention i WHERE voided=0 AND i.sub_county_id = %s AND i.implementing_partner_id IN %s """

        multiple_ip_ward_query = """select
  i.client_id, i.dreams_id, CONCAT_WS(" ",i.first_name, i.middle_name, i.last_name) AS client_name, i.date_of_birth, i.implementing_partner,  i.implementing_partner_id,i.county_of_residence,i.sub_county,
  i.ward, i.village, i.date_of_enrollment as date_of_enrollment, DATE(i.intervention_date) date_of_intervention, DATE(i.date_created) date_created, i.intervention as intervention_type, i.intervention_category, i.hts_result,
  i.pregnancy_test_result, i.client_ccc_number, i.date_linked_to_ccc,
  i.no_of_sessions_attended, i.comment, i.current_age, i.age_at_intervention
from stag_client_intervention i WHERE voided=0 AND i.ward_id = %s AND i.implementing_partner_id IN %s """


        multiple_ip_default_query = """select
  i.client_id, i.dreams_id, CONCAT_WS(" ",i.first_name, i.middle_name, i.last_name) AS client_name, i.date_of_birth, i.implementing_partner,  i.implementing_partner_id,i.county_of_residence,i.sub_county,
  i.ward, i.village, i.date_of_enrollment as date_of_enrollment, DATE(i.intervention_date) date_of_intervention, DATE(i.date_created) date_created, i.intervention as intervention_type, i.intervention_category, i.hts_result,
  i.pregnancy_test_result, i.client_ccc_number, i.date_linked_to_ccc,
  i.no_of_sessions_attended, i.comment, i.current_age, i.age_at_intervention
from stag_client_intervention i WHERE voided=0 AND i.implementing_partner_id IN %s """

        single_ip_sub_county_query = """select
  i.client_id, i.dreams_id, CONCAT_WS(" ",i.first_name, i.middle_name, i.last_name) AS client_name, i.date_of_birth, i.implementing_partner,  i.implementing_partner_id,i.county_of_residence,i.sub_county,
  i.ward, i.village, i.date_of_enrollment as date_of_enrollment, DATE(i.intervention_date) date_of_intervention, DATE(i.date_created) date_created, i.intervention as intervention_type, i.intervention_category, i.hts_result,
  i.pregnancy_test_result, i.client_ccc_number, i.date_linked_to_ccc,
  i.no_of_sessions_attended, i.comment, i.current_age, i.age_at_intervention
from stag_client_intervention i WHERE voided=0 AND i.sub_county_id = %s AND i.implementing_partner_id = %s """


        single_ip_ward_query = """select
  i.client_id, i.dreams_id, CONCAT_WS(" ",i.first_name, i.middle_name, i.last_name) AS client_name, i.date_of_birth, i.implementing_partner,  i.implementing_partner_id,i.county_of_residence,i.sub_county,
  i.ward, i.village, i.date_of_enrollment as date_of_enrollment, DATE(i.intervention_date) date_of_intervention, DATE(i.date_created) date_created, i.intervention as intervention_type, i.intervention_category, i.hts_result,
  i.pregnancy_test_result, i.client_ccc_number, i.date_linked_to_ccc,
  i.no_of_sessions_attended, i.comment, i.current_age, i.age_at_intervention
from stag_client_intervention i WHERE voided=0 AND i.ward_id = %s AND i.implementing_partner_id = %s """


        single_ip_default_query = """select
  i.client_id, i.dreams_id, CONCAT_WS(" ",i.first_name, i.middle_name, i.last_name) AS client_name, i.date_of_birth, i.implementing_partner,  i.implementing_partner_id,i.county_of_residence,i.sub_county,
  i.ward, i.village, i.date_of_enrollment as date_of_enrollment, DATE(i.intervention_date) date_of_intervention, DATE(i.date_created) date_created, i.intervention as intervention_type, i.intervention_category, i.hts_result,
  i.pregnancy_test_result, i.client_ccc_number, i.date_linked_to_ccc,
  i.no_of_sessions_attended, i.comment, i.current_age, i.age_at_intervention
from stag_client_intervention i
WHERE voided=0 AND i.implementing_partner_id = %s """

        try:

            ip_tuple_l = ip_list_str
            if sub_county is not None and sub_county:
                sub_county = int(sub_county)

            if ward is not None and ward:
                ward = int(ward)

            if len(ip_tuple_l) > 1:
                ip_list = tuple(ip_tuple_l)

                if ward is not None and ward:
                    cursor.execute(multiple_ip_ward_query, [ward, ip_list])
                elif sub_county is not None and sub_county:
                    cursor.execute(
                        multiple_ip_sub_county_query, [sub_county, ip_list])
                else:
                    cursor.execute(multiple_ip_default_query, [ip_list])
            else:
                ip_list = ip_list_str[0]
                if ward is not None and ward:
                    cursor.execute(single_ip_ward_query, [ward, ip_list])
                elif sub_county is not None and sub_county:
                    cursor.execute(single_ip_sub_county_query, [sub_county, ip_list])
                else:
                    cursor.execute(single_ip_default_query, [ip_list])

            print "Query was successful"
            columns = [col[0] for col in cursor.description]
            return [
                dict(zip(columns, row))
                for row in cursor.fetchall()
                ]
        except Exception as e:
            print 'There was an Error running the query: {} \n'.format(e)
            traceback.format_exc()

        return

    # add code to fetch line list of girls service layering

    def extract_service_layering_for_all_girls(self, ip_list_str, sub_county, ward):

        cursor = connection.cursor()
        multiple_ip_sub_county_query = "SELECT * FROM stag_individual_client_service_layering WHERE sub_county_id = %s AND  implementing_partner_id IN %s "
        multiple_ip_ward_query = "SELECT * FROM stag_individual_client_service_layering WHERE ward_id = %s AND  implementing_partner_id IN %s "
        multiple_ip_default_query = "SELECT * FROM stag_individual_client_service_layering WHERE implementing_partner_id IN %s "

        single_ip_sub_county_query = "SELECT * FROM stag_individual_client_service_layering WHERE sub_county_id = %s AND implementing_partner_id = %s "
        single_ip_ward_query = "SELECT * FROM stag_individual_client_service_layering WHERE ward_id = %s AND implementing_partner_id = %s "
        single_ip_default_query = "SELECT * FROM stag_individual_client_service_layering WHERE implementing_partner_id = %s "

        try:

            ip_tuple_l = ip_list_str
            if sub_county is not None and sub_county:
                sub_county = int(sub_county)

            if ward is not None and ward:
                ward = int(ward)

            if len(ip_tuple_l) > 1:
                ip_list = tuple(ip_tuple_l)

                if ward is not None and ward:
                    cursor.execute(multiple_ip_ward_query, [ward, ip_list])
                elif sub_county is not None and sub_county:
                    cursor.execute(
                        multiple_ip_sub_county_query, [sub_county, ip_list])
                else:
                    cursor.execute(multiple_ip_default_query, [ip_list])
            else:
                ip_list = ip_list_str[0]
                if ward is not None and ward:
                    cursor.execute(single_ip_ward_query, [ward, ip_list])
                elif sub_county is not None and sub_county:
                    cursor.execute(single_ip_sub_county_query, [sub_county, ip_list])
                else:
                    cursor.execute(single_ip_default_query, [ip_list])

            print "Query for individual service layering data was successful"
            columns = [col[0] for col in cursor.description]
            return [
                dict(zip(columns, row))
                for row in cursor.fetchall()
                ]
        except Exception as e:
            print 'There was an Error running query for individual service layering data\n'
            traceback.format_exc()

        return


    def load_workbook(self):
        DREAMS_TEMPLATE_PLAIN = os.path.join(settings.BASE_DIR, '../templates/excel_template/dreams_export.xlsx')
        try:
            wb = xl.load_workbook(DREAMS_TEMPLATE_PLAIN)
            return wb
        except InvalidFileException as e:
            traceback.format_exc()

    def load_intervention_workbook(self):
        DREAMS_INTERVENTION_TEMPLATE = os.path.join(settings.BASE_DIR, '../templates/excel_template/service_uptake_template.xlsx')
        try:
            wb = xl.load_workbook(DREAMS_INTERVENTION_TEMPLATE)
            return wb
        except InvalidFileException as e:
            traceback.format_exc()

    def load_service_layering_workbook(self):
        DREAMS_SERVICE_LAYERING_TEMPLATE = os.path.join(settings.BASE_DIR, '../templates/excel_template/individual_service_layering_template.xlsx')
        try:
            wb = xl.load_workbook(DREAMS_SERVICE_LAYERING_TEMPLATE)
            return wb
        except InvalidFileException as e:
            traceback.format_exc()

    def prepare_excel_doc(self, response, ip_list_str, sub_county, ward, show_PHI):

        try:
            print "Starting DB Query! ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db_data, cursor = self.get_export_rows(ip_list_str, sub_county, ward)

            writer = csv.writer(response, quoting=csv.QUOTE_ALL)
            col_names = []

            for col in cursor.description:
                col_names.append(col[0])
            writer.writerow(col_names)

            for row in cursor:
                writer.writerow(row)

            print "Finished DB Query. Rendering Now. ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print "Completed rendering excel ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        except InvalidFileException as e:
            traceback.format_exc()
        except ReadOnlyWorkbookException as e:
            traceback.format_exc()
        except SheetTitleException as e:
            traceback.format_exc()
        except Exception as e:
            traceback.format_exc()
        return

    def get_intervention_excel_doc(self, ip_list_str, sub_county, ward, show_PHI):

        try:

            wb = self.load_intervention_workbook()
            interventions_sheet = wb.get_sheet_by_name('DREAMS_Services')
            print "Starting Intervention DB Query! ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db_data = self.fetch_intervention_rows(ip_list_str, sub_county, ward)
            print "Finished Intervention DB Query. Rendering Now. ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            i = 1
            for row in db_data:
                i += 1
                self.map_interventions(interventions_sheet, i, row, show_PHI)

            wb.save('dreams_interventions.xlsx')
            print "Completed rendering excel ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            return wb
        except InvalidFileException as e:
            traceback.format_exc()
        except ReadOnlyWorkbookException as e:
            traceback.format_exc()

        except SheetTitleException as e:
            traceback.format_exc()
        return

    # procedure for handling individual service layering report
    def get_individual_layering_report(self, ip_list_str, sub_county, ward, show_PHI):

        try:

            wb = self.load_service_layering_workbook()
            main_sheet = wb.get_sheet_by_name('Services_Received')
            print "Starting Query for individual service layering report! ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db_data = self.extract_service_layering_for_all_girls(ip_list_str, sub_county, ward)
            print "Finished Query for individual service layering report!. Rendering Now. ", datetime.datetime.now().strftime(
                '%Y-%m-%d %H:%M:%S')
            i = 1
            for row in db_data:
                i += 1
                self.map_individual_service_layering_report(main_sheet, i, row, show_PHI)

            wb.save('Dreams_individual_service_layering_report.xlsx')
            print "Completed rendering Individual Service Layering Report ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            return wb
        except InvalidFileException as e:
            traceback.format_exc()
        except ReadOnlyWorkbookException as e:
            traceback.format_exc()

        except SheetTitleException as e:
            traceback.format_exc()
        return

    # Mapping for individual service layering report
    def map_individual_service_layering_report(self, ws, i, row, show_PHI):

        phi_cols = {
            'first_name': 2,
            'middle_name': 3,
            'last_name': 4
        }

        open_access_cols = {
            'dreams_id': 1,
            'date_of_birth': 5,
            'date_of_enrollment': 6,
            'age_at_enrollment': 7,
            'current_age': 8,
            'implementing_partner': 9,
            'county_of_residence': 10,
            'sub_county': 11,
            'ward': 12,
            'village': 13,
            'shuga_II': 14,
            'respect_k': 15,
            'hcbf': 16,
            'mhmc': 17,
            'sister_to_sister_k': 18,
            'mlrc': 19,
            'behavioral_other': 20,
            'hts_client': 21,
            'hts_partner': 22,
            'linkage_to_ccc': 25,
            'pregnancy_test': 27,
            'anc_pmtct': 29,
            'sti_screening': 30,
            'sti_treatment': 31,
            'sti_linkage': 32,
            'tb_screening': 33,
            'linked_for_tb_treatment': 34,
            'condom_education_and_demo': 35,
            'condom_provided': 36,
            'partner_vmmc': 37,
            'contraception_education': 38,
            'contraception_ind_counseling': 39,
            'contraception_pills_oral': 40,
            'contraception_injectable': 41,
            'contraception_implant': 42,
            'contraception_iud_coil': 43,
            'prep_given': 44,
            'prep_education': 45,
            'sexual_violence_pep': 46,
            'sexual_violence_pss': 47,
            'sexual_violence_rescue_shelter': 48,
            'sexual_violence_police': 49,
            'sexual_violence_trauma_counseling': 50,
            'sexual_violence_emergency_contraception': 51,
            'sexual_violence_exam_treatment': 52,
            'education_school_fees': 62,
            'education_stationery': 63,
            'education_uniform': 64,
            'education_other_support': 65,
            'parent_program_fmp': 79,
            'economic_strengthening_fc_training': 66,
            'economic_strengthening_voc_training': 67,
            'economic_strengthening_microfinance': 68,
            'economic_strengthening_internship': 69,
            'economic_strengthening_startups': 70,
            'cash_transfer': 74,
            'ovc_for_children_sibling_other': 75,
            'nutritional_support': 76,
            'drug_addiction_counseling': 77,
            'sab': 78,
            'hts_client_linked_to_hts': 23,
            'pregnancy_test_confirmed_linkage': 28,
            'hts_partner_linked_to_hts': 24,
            'positive_partner_linked_to_ccc': 26,
            'tube_ligation': 81,
            'sexual_violence_legal_support': 53,
            'economic_strengthening_employment': 71,
            'economic_strengthening_entrep_training': 72,
            'economic_strengthening_entrep_support': 73,
            'sexual_violence_other': 54,
            'physical_violence_pss': 55,
            'physical_violence_rescue_shelter': 56,
            'physical_violence_police': 57,
            'physical_violence_trauma_counseling': 58,
            'physical_violence_exam_treatment': 59,
            'physical_violence_legal_support': 60,
            'physical_violence_other': 61,
            'parent_program_fmp2': 80,
            'bio_medical_other': 82,
            'social_protection_other': 83,
            'received_behavioral_interventions': 84,
            'received_biomedical_interventions': 85,
            'received_post_gbv_interventions': 86,
            'received_social_protection_interventions': 87,
            'received_other_interventions': 88,
            'duration_in_dreams_program': 89,
            'exited_from_program': 90,
            'date_exited': 91
        }

        # Hide PHI column values where necessary

        if show_PHI:
            cols = open_access_cols.copy()
            cols.update(phi_cols)
        else:
            cols = open_access_cols

        for k, v in cols.items():
            ws.cell(row=i, column=v, value=row.get(k))

    def map_interventions(self, ws, i, row, show_PHI):

        phi_cols = {
            'client_name': 3
        }

        open_access_cols = {
            'dreams_id': 2,
            'implementing_partner': 4,
            'date_of_birth': 5,
            'date_of_enrollment': 6,
            'county_of_residence': 7,
            'sub_county': 8,
            'ward': 9,
            'village': 10,
            'date_of_intervention': 11,
            'age_at_intervention': 12,
            'current_age': 13,
            'intervention_type': 14,
            'intervention_category': 15,
            'hts_result': 16,
            'pregnancy_test_result': 17,
            'client_ccc_number': 18,
            'date_linked_to_ccc': 19,
            'no_of_sessions_attended': 20,
            'comment': 21,
        }

        # Hide PHI column values where necessary

        if show_PHI:
            cols = open_access_cols.copy()
            cols.update(phi_cols)
        else:
            cols = open_access_cols

        for k, v in cols.items():
            ws.cell(row=i, column=v, value=row.get(k))

    def map_demographics(self, ws, i, row, show_PHI):

        # These are columns that contain identifiers. They should be
        phi_cols = {
            'first_name': 4,
            'middle_name': 5,
            'last_name': 6,
            'verification_doc_no': 10,
            'phone_number': 17,
            'dss_id_number': 26,
            'guardian_name': 27,
            'guardian_phone_number': 32,
            'guardian_national_id': 33
        }

        open_access_cols = {
            'implementing_partner': 2,
            'IP_Code': 3,
            'date_of_birth': 7,
            'verification_document': 8,
            'verification_document_other': 9,
            'date_of_enrollment': 11,
            'current_age': 13,
            'age_at_enrollment': 14,
            'marital_status': 16,
            'county_name': 18,
            'sub_county_name': 19,
            'ward_id': 21,
            'ward_name': 20,
            'informal_settlement': 22,
            'village': 23,
            'landmark': 24,
            'dreams_id': 25,
            'relationship_with_guardian': 30,
            'exit_status': 204,
            'exit_date': 205,
            'exit_reason': 206
        }

        # merge PIH and open access columns based on permissions

        if show_PHI:
            cols = open_access_cols.copy()
            cols.update(phi_cols)
        else:
            cols = open_access_cols

        for k, v in cols.items():
            if k == 'IP_Code':
                val = row.get('implementing_partner_id')
                ws.cell(row=i, column=v, value=val)
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_individual_and_household(self, ws, i, row):
        cols = {
            'head_of_household': 34,
            'head_of_household_other': 35,
            'age_of_household_head': 36,
            'father_alive': 37,
            'mother_alive': 38,
            'parent_chronically_ill': 39,
            'main_floor_material': 40,
            'main_floor_material_other': 41,
            'main_roof_material': 42,
            'main_roof_material_other': 43,
            'main_wall_material': 44,
            'main_wall_material_other': 45,
            'source_of_drinking_water': 46,
            'source_of_drinking_water_other': 47,
            'no_of_people_in_household': 57,
            'no_of_adults': 60,
            'no_of_females': 58,
            'no_of_males': 59,
            'no_of_children': 61,
            'currently_in_ct_program': 63,
            'current_ct_program': 64,
            'ever_enrolled_in_ct_program': 62,
            'ever_missed_full_day_food_in_4wks': 48,
            'has_disability': 50,
            'no_of_days_missed_food_in_4wks': 49,
            'disability_types': 51
        }

        for k, v in cols.items():
            if k == 'disability_types':
                val = row.get(k)
                if val is not None:
                    dts = val.split(",")
                    for d in dts:
                        if d == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_disability_type().get(d), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_sexuality(self, ws, i, row):
        cols = {
            'age_at_first_sexual_encounter': 108,
            'sex_partners_in_last_12months': 109,
            'age_of_last_partner': 111,
            'age_of_second_last_partner': 112,
            'age_of_third_last_partner': 113,
            'ever_had_sex': 107,
            'has_sexual_partner': 110,
            'know_last_partner_hiv_status': 117,
            'know_second_last_partner_hiv_status': 118,
            'know_third_last_partner_hiv_status': 119,
            'last_partner_circumcised': 114,
            'received_money_gift_for_sex': 123,
            'second_last_partner_circumcised': 115,
            'third_last_partner_circumcised': 116,
            'used_condom_with_last_partner': 120,
            'used_condom_with_second_last_partner': 121,
            'used_condom_with_third_last_partner': 122
        }
        for k, v in cols.items():
            ws.cell(row=i, column=v, value=row.get(k))

    def map_reproductive_health(self, ws, i, row):
        cols = {
            'no_of_biological_children': 125,
            'anc_facility_name': 128,
            'known_fp_method_other': 134,
            'current_fp_method_other': 139,
            'reason_not_using_fp_other': 141,
            'current_anc_enrollment': 127,
            'current_fp_method': 138,
            'currently_pregnant': 126,
            'currently_use_modern_fp': 137,
            'fp_methods_awareness': 129,
            'has_biological_children': 124,
            'reason_not_using_fp': 140,
            'known_fp_methods': 130
        }
        for k, v in cols.items():
            if k == 'known_fp_methods':
                val = row.get(k)
                if val is not None:
                    fpm = val.split(",")
                    for m in fpm:
                        if m == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_fp_method().get(m), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_drug_use(self, ws, i, row):
        cols = {
            # 'dr.drug_abuse_last_12months_other': 180,
            'drug_used_last_12months_other': 191,
            'drug_abuse_last_12months': 184,
            'frequency_of_alcohol_last_12months': 183,
            'produced_alcohol_last_12months': 194,
            'used_alcohol_last_12months': 182,
            'drugs_used_in_last_12_months': 185
        }
        for k, v in cols.items():
            if k == 'drugs_used_in_last_12_months':
                val = row.get(k)
                if val is not None:
                    drugs = val.split(",")
                    for d in drugs:
                        if d == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_drugs().get(d), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_program_participation(self, ws, i, row):
        cols = {
            'dreams_program_other': 203,
            'programmes_enrolled': 195
        }
        for k, v in cols.items():
            if k == 'programmes_enrolled':
                val = row.get(k)
                if val is not None:
                    pgs = val.split(",")
                    for p in pgs:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_dreams_programmes().get(p), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_gbv(self, ws, i, row):
        cols = {
            'gbv_help_provider_other': 169,
            'preferred_gbv_help_provider_other': 181,
            'economic_threat_ever': 148,
            'economic_threat_last_3months': 149,
            'humiliated_ever': 142,
            'humiliated_last_3months': 143,
            'insulted_ever': 146,
            'insulted_last_3months': 147,
            'knowledge_of_gbv_help_centres': 170,
            'physical_violence_ever': 150,
            'physical_violence_last_3months': 151,
            'physically_forced_other_sex_acts_ever': 154,
            'physically_forced_other_sex_acts_last_3months': 155,
            'physically_forced_sex_ever': 152,
            'physically_forced_sex_last_3months': 153,
            'seek_help_after_gbv': 158,
            'threatened_for_sexual_acts_ever': 156,
            'threatened_for_sexual_acts_last_3months': 157,
            'threats_to_hurt_ever': 144,
            'threats_to_hurt_last_3months': 145,
            'providers_sought': 159,
            'preferred_providers': 171,
        }
        for k, v in cols.items():
            if k == 'providers_sought':
                val = row.get(k)
                if val is not None:
                    sp = val.split(",")
                    for p in sp:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_gbv_sought_provider().get(p), value='Yes')
            elif k == 'preferred_providers':
                val = row.get(k)
                if val is not None:
                    sp = val.split(",")
                    for p in sp:
                        if p == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_gbv_preferred_provider().get(p), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_education_and_employment(self, ws, i, row, show_PHI):

        # These are columns that contain identifiers. They should be
        phi_cols = {
            'current_school_name': 66
        }

        open_access_cols = {
            'current_class': 70,
            'current_school_level_other': 69,
            'current_edu_supporter_list': 71,
            'current_education_supporter_other': 76,
            'reason_not_in_school_other': 78,
            'dropout_class': 80,
            'life_wish_other': 83,
            'current_income_source_other': 85,
            'banking_place_other': 88,
            'banking_place': 87,
            'current_income_source': 84,
            'current_school_level': 68,
            'current_school_type': 67,
            'currently_in_school': 65,
            'dropout_school_level': 81,
            'has_savings': 86,
            'last_time_in_school': 79,
            'life_wish': 82,
            'reason_not_in_school': 77
        }

        # merge data base on privileges

        if show_PHI:
            cols = open_access_cols.copy()
            cols.update(phi_cols)
        else:
            cols = open_access_cols

        for k, v in cols.items():
            if k == 'current_edu_supporter_list':
                val = row.get(k)
                if val is not None:
                    sps = val.split(",")
                    for s in sps:
                        if s == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_education_supporter().get(s), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_hiv_testing(self, ws, i, row):
        cols = {
            'care_facility_enrolled': 93,
            'reason_not_in_hiv_care_other': 95,
            'reason_not_tested_for_hiv': 96,
            'reason_never_tested_for_hiv_other': 103,
            'enrolled_in_hiv_care': 92,
            'ever_tested_for_hiv': 89,
            'knowledge_of_hiv_test_centres': 106,
            'last_test_result': 91,
            'period_last_tested': 90,
            'reason_not_in_hiv_care': 94
        }
        for k, v in cols.items():
            if k == 'reason_not_tested_for_hiv':
                val = row.get(k)
                if val is not None:
                    rns = val.split(",")
                    for r in rns:
                        if r == 96:
                            continue
                        else:
                            ws.cell(row=i, column=self.map_reason_never_tested_for_hiv().get(r), value='Yes')
            else:
                ws.cell(row=i, column=v, value=row.get(k))

    def map_dreams_programmes(self):
        return {
            '1': 195,
            '2': 196,
            '3': 197,
            '4': 198,
            '5': 199,
            '6': 200,
            '7': 201,
            '8': 202,
            '96': 203
        }

    def map_drugs(self):
        return {
            '1': 185,
            '2': 186,
            '3': 187,
            '4': 188,
            '5': 189,
            '6': 190,
            '7': 191,
            '8': 192,
            '96': 193

        }

    def map_gbv_preferred_provider(self):
        return {
            '1': 171,
            '2': 172,
            '3': 173,
            '4': 174,
            '5': 175,
            '6': 176,
            '7': 177,
            '8': 178,
            '9': 179,
            '10': 180,
            '96': 181

        }

    def map_gbv_sought_provider(self):
        return {
            '1': 159,
            '2': 160,
            '3': 161,
            '4': 162,
            '5': 163,
            '6': 164,
            '7': 165,
            '8': 166,
            '9': 167,
            '10': 168,
            '96': 169

        }

    def map_fp_method(self):
        return {
            '1': 130,
            '2': 131,
            '3': 132,
            '4': 133,
            '5': 134,
            '6': 135,
            '96': 136

        }

    def map_reason_never_tested_for_hiv(self):
        return {
            '1': 96,
            '2': 97,
            '3': 98,
            '4': 99,
            '5': 100,
            '6': 101,
            '7': 102,
            '8': 103,
            '9': 104,
            '96': 105

        }

    def map_education_supporter(self):
        return {
            '1': 71,
            '2': 72,
            '3': 73,
            '4': 74,
            '5': 75,
            '96': 76
        }

    def map_disability_type(self):
        return {
            '1': 51,
            '2': 52,
            '3': 53,
            '4': 54,
            '5': 55,
            '96': 56
        }

    def map_implementing_partner(self):
        return {
            1: 'Afya Jijini',
            2: 'AIHA',
            3: 'Aphiaplus Western',
            4: 'Global Communities',
            5: 'Henry Jackson Foundation',
            6: 'HWWK',
            7: 'IMC',
            8: 'IRDO',
            9: 'LVCT Health',
            10: 'Peace Corps'
        }

    def map_verification_document(self):
        return {
            4: 'Pending Verification',
            1: 'Birth Certificate',
            2: 'National ID',
            3: 'National ID waiting card',
            96: 'Other (Specify)',
        }

    def map_marital_status_codes(self):
        return {
            1: 'Single',
            2: 'Married / Cohabiting',
            3: 'Separated / Divorced',
            4: 'Widowed',
        }




    def headOfHouseHoldDictionary(self):
        return {
            1: "Self",
            2: "Father",
            3: "Mother",
            4: "Sibling",
            5: "Uncle/Aunt",
            6: "Grandparents",
            7: "Husband/Partner",
            96: "Other/Specify"
        }

    def yesNoDictionary(self):
        return {
            1: "YES",
            2: "NO",
            3: "Don't Know",
            4: "Not Applicable"
        }

    def floorMaterialDictionary(self):
        return {
            1: "Earth/Mud/Dung/Sand",
            2: "Wood planks",
            3: "Ceramic tiles",
            4: "Cement",
            96: "Other (Specify)"
        }

    def wallMaterialDictionary(self):
        return {
            1: "No Walls",
            2: "Dung/Mud",
            3: "Stone with mud",
            4: "Plywood/Cardboard",
            5: "Carton",
            6: "Wood",
            7: "Stone/Cement",
            96: "Other (Specify)"
        }

    def roofMaterialDictionary(self):
        return {
            1: "Grass/Thatch/Makuti",
            2: "Tin cans",
            3: "Corrugated iron sheet",
            4: "Asbestos sheets",
            5: "Concrete",
            96: "Other (Specify)"
        }

    def drinkingWaterSourceDictionary(self):
        return {
            1: "Piped water",
            2: "Open well",
            3: "Covered well/borehole",
            4: "Surface water (River, Spring and Lakes)",
            5: "Rain water",
            96: "Other (Specify)"
        }

    def frequencyDictionary(self):
        return {
                    1: "Rarely (1-2 days)",
            2: "Sometimes (3-10 days)",
            3: "Often (more than 10 days)",
            4: "Always",
            5: "Sometimes",
            6: "Never",
            7: "Often",
            8: "Not in the last 3 months",
            9: "Everyday",
            10: "5 to 6 times a week",
            11: "3 to 4 times a week",
            12: "Once a week",
            13: "2 to 3 times a month",
            14: "Once a month",
            15: "3 to 11 times in the past year",
            16: "1 to 2 times in the past year",
            96: "Other (Specify)"

        }

    def schoolTypeDictionary(self):
        return {
            1: "Formal",
            2: "Informal"
        }

    def schoolLevelDictionary(self):
        return {
            1: "Nursery",
            2: "PRIMARY_LEVEL",
            3: "SECONDAY_LEVEL",
            4: "TERTIARY_LEVEL",
            5: "VOCATIONAL_LEVEL",
            96: "Other (Specify)"
        }


    def reasonNotInSchoolDictionary(self):
        return {
            1: "Completed High School",
            2: "Lack of fees",
            3: "Pregnancy",
            4: "Peer pressure",
            5: "Not interested",
            6: "Awaiting to join secondary",
            7: "Got married",
            96: "Other (Specify)"
        }

    def lastInSchoolDictionary(self):
        return {
            1: "Less than 6 months ago",
            2: "6 to 12 months",
            3: "1to 2 years ago",
            4: "More than 2 years",
            5: "Never attended school",
            96: "Other (Specify)"
        }

    def lifeWishDictionary(self):
        return {
            1: "Pursue a course",
            2: "Start a business",
            3: "Go back to school",
            4: "Get married",
            96: "Other (Specify)"
        }

    def incomeSourceDictionary(self):
        return {
            1: "Formally employed",
            2: "Business person",
            3: "Casual labour",
            4: "Petty trade",
            5: "Farmer",
            6: "None",
            96: "Other (Specify)"
        }

    def bankingPlaceDictionary(self):
        return {
            1: "At home",
            2: "Table banking",
            3: "In the bank",
            96: "Other (Specify)"
        }

    def periodDictionary(self):
        return {
            1: "Less than 6 Months ago",
            2: "6 - 12 months ago",
            3: "1- 2 years ago",
            4: "More than 2 years ago",
            5: "Never attended to school",
            6: "Less than 3 months ago",
            7: "3-5 months ago",
            8: "6-12 months ago",
            9: "More than 12 months ago"
        }

    def hivTestDictionary(self):
        return {
            1: "Positive",
            2: "Negative",
            3: "Don't Know",
            4: "Declined to disclose"
        }

    def reasonNotInCareDictionary(self):
        return {
            1: "Facility is too far away",
            2: "I don't know where clinic is",
            3: "I can't afford it",
            4: "I feel healthy/not sick",
            5: "I fear people will know I have HIV if I go to clinic",
            6: "I feel I will be discriminated against by people at a facility",
            7: "The providers at facility are unfriendly",
            8: "I am taking alternative medicine that is not availble at the clinic",
            9: "I,m too busy to go",
            96: "Other(specify)"
        }

    def relativeAgeDictionary(self):
        return {
            2: "Same Age",
            3: "Younger",
            4: "Older"
        }

    def familyPlanningDictionary(self):
        return {
            1: "Pills",
            2: "Injectable",
            3: "Implants",
            4: "IUCD",
            5: "Condom",
            6: "Permanent (Tube Ligation)",
            96: "Other (Specify)"
        }

    def reasonNoInFPDictionary(self):
        return {
           1: "Not sexually active",
           2: "Religious reasons",
           3: "Can not afford",
           4: "Do not know where to get",
           5: "Currently pregnant",
           96: "Other (Specify)"
        }

    def get_intervention_excel_transferred_in_doc(self, ip, from_intervention_date, to_intervention_date, show_PHI):

        try:

            wb = self.load_intervention_workbook()
            interventions_sheet = wb.get_sheet_by_name('DREAMS_Services')
            print "Starting Intervention DB Query! ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            db_data = self.fetch_intervention_transferred_in_rows(ip, from_intervention_date, to_intervention_date)
            print "Finished Intervention DB Query. Rendering Now. ", datetime.datetime.now().strftime(
                '%Y-%m-%d %H:%M:%S')
            i = 1
            for row in db_data:
                i += 1
                self.map_interventions(interventions_sheet, i, row, show_PHI)

            wb.save('dreams_interventions.xlsx')
            print "Completed rendering excel ", datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            return wb
        except (InvalidFileException, ReadOnlyWorkbookException, SheetTitleException) as e:
            traceback.format_exc()
        return

    def fetch_intervention_transferred_in_rows(self, ip, from_intervention_date, to_intervention_date):
        cursor = connection.cursor()

        query = """select
          i.client_id, i.dreams_id, CONCAT_WS(" ",i.first_name, i.middle_name, i.last_name) AS client_name, i.date_of_birth, 
          i.implementing_partner,  i.implementing_partner_id,i.county_of_residence,i.sub_county,
          i.ward, i.village, i.date_of_enrollment as date_of_enrollment, DATE(i.intervention_date) date_of_intervention, 
          DATE(i.date_created) date_created, i.intervention as intervention_type, i.intervention_category, i.hts_result,
          i.pregnancy_test_result, i.client_ccc_number, i.date_linked_to_ccc,
          i.no_of_sessions_attended, i.comment, i.current_age, i.age_at_intervention
          from stag_client_intervention i
          WHERE i.voided=0 AND i.transferred_client=1 
          """
        params = []

        if ip is not None:
            query += " AND i.implementing_partner_id = %s "
            params.append(ip.id)
        if from_intervention_date is not None and from_intervention_date:
            query += " AND i.intervention_date >= %s "
            params.append(from_intervention_date)
        if to_intervention_date is not None and to_intervention_date:
            query += " AND i.intervention_date <= %s "
            params.append(to_intervention_date)

        try:
            cursor.execute(query, params)
            print "Query was successful"
            columns = [col[0] for col in cursor.description]
            return [
                dict(zip(columns, row))
                for row in cursor.fetchall()
            ]
        except Exception as e:
            print 'There was an Error running the query: {} \n'.format(e)
            traceback.format_exc()

        return
